from openpyxl import load_workbook , Workbook
import shutil
import sys

def dealwith_excel(xls_file:str):
    # "UART_final_202301010.xls"
    wb = load_workbook(xls_file)
    sheetNames = wb.sheetnames
    for sh_name in sheetNames:
        if sh_name in ('preface', 'module_tpl','device'):
            continue
        else:
            out_file = './split/'+sh_name+'.xlsx'
            shutil.copyfile(xls_file, out_file)
            out_wb  = load_workbook(out_file)
            out_sheets = out_wb.sheetnames
            for sh in out_sheets:
                if sh != sh_name:
                    out_wb.remove(out_wb[sh])
                    pass
            out_wb.save(out_file)
            pass
    # 再生成一个device
    out_file = './split/device_base.xlsx'
    shutil.copyfile(xls_file, out_file)
    out_wb  = load_workbook(out_file)
    for sh_name in sheetNames:
        if sh_name not in ('preface', 'module_tpl','device'):
            out_wb.remove(out_wb[sh])
            pass
        pass
    out_wb.save(out_file)

# fieldRWOp_arrUpper = ('RO', 'RW', 'RC', 'RS', 'WRC', 'WRS', 'WC', 'WS', 'WSRC', 'WCRS', 'W1C', 'W1S', 'W1T', 'W0C', 'W0S', 'W0T',
#                  'W1SRC', 'W1CRS', 'W0SRC', 'W0CRS', 'WO', 'WOC', 'WOS', 'W1', 'WO1')
    
def printAllValues(strlst: list):
    # strop =''
    # bFirst = True
    # for op in strlst:
    #     if not bFirst:
    #         strop += ','
    #     else:
    #         bFirst = False
    #     strop += op
    # print(strop)
    pass


if __name__ == '__main__':
    # 全路径是为方便在vscode中进行调试
    file_name = './xy2_mp32daptyxx_DDF 240118.xlsx'
    if len(sys.argv) == 2:
        file_name= sys.argv[1] 
    dealwith_excel(file_name)
