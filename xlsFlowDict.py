import os,sys
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet


def checkDeviceSheet(ws:worksheet):
    pass

def getValueFromModuleDict(attrName: str, moduleDict : dict):
    val = None
    if attrName in moduleDict:
        val = moduleDict[attrName]
        pass
    return val



def checkModuleSheet(ws:worksheet):
    sheettitle = ws['A1'].value
    moduleHeader_dict ={}
    module_lst =[]
    if sheettitle == 'peripheral:':
        #还是需要先读取所有flag的row 行号
        nRows = ws.max_row
        flag_names = ('addressBlocks:', 'interrupts:',
                    'cluster:', 'end cluster', 'register:')
        flag_dict = {}
        flag_rows = []
        emptyA_rows = []
        row = 3
        while row <= nRows:
            val = ws.cell(row, 1).value
            if val in flag_names:
                flag_rows.append(row)
                if val in flag_dict:
                    flag_dict[val].append(row)
                else:
                    flag_dict[val] = [row]
                pass
            elif not val:
                bEmpty = True
                for col in range(2,6):
                    val = ws.cell(row, col).value
                    if val:
                        bEmpty = False
                        break
                if bEmpty:
                    emptyA_rows.append(row)
            row += 1
        row_end = nRows
        if flag_rows:
            row_end = flag_rows[0]

        rowb = ws[2]
        for cellitem in rowb:
            if cellitem.value is not None:
                # print(f'col: {cellitem.column}, value: {cellitem.value}')
                moduleHeader_dict[cellitem.column] = cellitem.value
                pass
            pass
        for row_i in range(3,row_end):
            rowb = ws[row_i]
            bEmptyRow = True
            module_dict ={}
            for cellitem in rowb:
                if cellitem.value is not None:
                    bEmptyRow = False
                    pass
                if cellitem.column in moduleHeader_dict:
                    module_dict[moduleHeader_dict[cellitem.column]] = cellitem.value
                    pass
                pass
            pass
            if not bEmptyRow:
                module_lst.append(module_dict)
                pass
            else:
                break
        pass
        #读取register
        reg_lst= []
        regHeader_dict ={}
        reg_roos = flag_dict['register:']
        for reg_row in reg_roos:
            rowb = ws[reg_row+1]
            for cellitem in rowb:
                if cellitem.value is not None:
                    # print(f'col: {cellitem.column}, value: {cellitem.value}')
                    regHeader_dict[cellitem.column] = cellitem.value
                    pass
                pass
            for row_i in range(reg_row+2,reg_row+3):
                rowb = ws[row_i]
                bEmptyRow = True
                reg_dict = {}
                for cellitem in rowb:
                    if cellitem.value is not None:
                        bEmptyRow = False
                        pass
                    if cellitem.column in regHeader_dict:
                        reg_dict[regHeader_dict[cellitem.column]] = cellitem.value
                        pass
                    pass
                if not bEmptyRow:
                    reg_lst.append(reg_dict)
                pass
        for reg in reg_lst:
            print(reg)
        # 需要判断是否是cluster里的register
        cluster_rows = flag_dict['cluster:']
        end_cluster_rows = flag_dict['end cluster']
        # 先匹配cluster 和 end cluster  
        # 然后再针对具体的cluster end clu 对来读取
        # 这部分处理逻辑，建议参考./xlsFlowX.py中
    # for module in module_lst:
    #     print(module)
    pass

def dealwith_excel(xls_file:str, outFlag:int = 1):
    # "UART_final_202301010.xls"
    wb = load_workbook(xls_file, data_only=True)
    sheetNames = wb.sheetnames
    st_dev = None
    checkErr = False
    module_file_lst=[]
    module_sysDomain_Perip_dict ={}
    for sh_name in sheetNames:
        # 遍历 sheet, preface sheet不处理, module_tpl sheet不处理
        if sh_name in ('preface', 'module_tpl'):
            continue
        elif sh_name == 'device':
            # 先读取device sheet
            ws = wb[sh_name]
            checkDeviceSheet(ws)
            # print(st_dev.get_inst_str())
        else:
            # 再逐一读取peripheral sheet
            ws = wb[sh_name]
            checkModuleSheet(ws)
            pass
                 

    if checkErr:
        filename = os.path.basename(xls_file)
        # out_mark_xlsx_file = filename.replace('.xlsx', '_errMk.xlsx')
        print("Check Failed. Please review "+ filename + " and fix it.")
        # wb.save(out_mark_xlsx_file)
    else:
        pass


if __name__ == '__main__':
    # 全路径是为方便在vscode中进行调试
    # file_name = 'D:/workspace/demopy/excel_flow/excel/ahb_cfg_20230925.xlsx'
    file_name = './xy2_mp32daptyxx_DDF 240118.xlsx'
    dealwith_excel(file_name)