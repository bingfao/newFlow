#####################################################################
# use openpyxl to deal with excel file to export .h and .svh for DV
#####################################################################

# Author: binga.gao
# date: 2023-10-24
# change-desc:
#             change-date: 2023-10-24
#             1. change use xlrd to xlsx
#             2. show check error info in xlsx file with border color red
#             3. support virtual reg and  reg  group


import math
import re
from datetime import date
import os
import os,sys
import copy
from sys import flags
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
from openpyxl.styles import colors, Border, Side, Font, Color



# import xlrd


char_arr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
# busTypestr_arr = ("AHB", "AXI")
bitWidMask_arr = ('0x01', '0x03', '0x07', '0x0F', '0x1F', '0x3F', '0x7F', '0xFF', '0x01FF', '0x03FF', '0x07FF', '0x0FFF', '0x1FFF', '0x3FFF', '0x7FFF', '0xFFFF',
                  '0x01FFFF', '0x03FFFF', '0x07FFFF', '0x0FFFFF', '0x1FFFFF', '0x3FFFFF', '0x7FFFFF', '0xFFFFFF', '0x01FFFFFF', '0x03FFFFFF', '0x07FFFFFF', '0x0FFFFFFF', '0x1FFFFFFF', '0x3FFFFFFF', '0x7FFFFFFF', '0xFFFFFFFF',
                  '0x01FFFFFF', '0x03FFFFFF', '0x07FFFFFF', '0x0FFFFFFF', '0x1FFFFFFF', '0x3FFFFFFF', '0x7FFFFFFF', '0xFFFFFFFF', '0x01FFFFFF', '0x03FFFFFFFF', '0x07FFFFFFFF', '0x0FFFFFFFFF', '0x1FFFFFFFFF', '0x3FFFFFFFFF', '0x7FFFFFFFFF', '0xFFFFFFFFFF',
                  '0x01FFFFFFFF', '0x03FFFFFFFF', '0x07FFFFFFFF', '0x0FFFFFFFFF', '0x1FFFFFFFFF', '0x3FFFFFFFFF', '0x7FFFFFFFFF', '0xFFFFFFFFFF', '0x01FFFFFFFF', '0x03FFFFFFFFFF', '0x07FFFFFFFFFF', '0x0FFFFFFFFFFF', '0x1FFFFFFFFFFF', '0x3FFFFFFFFFFF', '0x7FFFFFFFFFFF', '0xFFFFFFFFFFFF')

fieldRWOp_arr = ('rw', 'ro', 'wo', 'w1', 'w1c', 'rc', 'rs', 'wrc', 'wrs', 'wc', 'ws', 'wsrc', 'wcrs', 'w1s', 'w1t', 'w0c',
                 'w0s', 'w0t', 'w1src', 'w1crs', 'w0src', 'w0crs', 'woc', 'wos', 'wo1')

cst_tab_str='    '  #用来替代\t的格式化用字符

accessDict = {'R': '__I ', 'W': '__O ', 'RW': '__IO'}

uint_dict = {8: 'uint8_t', 16: 'uint16_t',
                         32: 'uint32_t', 64: 'uint64_t'}

ubitSize_bytes_dict = {8: '1', 16: '2',
                         32: '4', 64: '8'}


cst_HEXValue_StringSize = 10
cst_emIRQ_SpaceSize = 20
cst_IRQ_SpaceSize = 40
cst_CORE_SpaceSize = 40
cst_Perip_SpaceSize = 40
cst_Mem_SpaceSize = 40
cst_Reg_SpaceSize = 40
cst_RegField_SpaceSize = 48

# RW, RO, WO, W1, W1C, RC, RS, WRC, WRS, WC, WS, WSRC, WCRS, W1S, W1T, W0C,
#                  W0S, W0T, W1SRC, W1CRS, W0SRC, W0CRS, W0C, W0S, WO1

# uint_type_arr = ('uint8_t', 'uint16_t', 'uint32_t', 'uint64_t')

def getBoolStr(val: bool):
    rt_str = 'false'
    if val:
        rt_str='true'
    return rt_str

def get_output_c_dir():
    if sys.platform == 'linux':
        prj_root = os.getenv("PRJ_ROOT")
        result_dir    = os.path.join(prj_root,'dv/tb/reg_model/c')
        return result_dir

def get_output_dut_cfg_dir():
    if sys.platform == 'linux':
        prj_root = os.getenv("PRJ_ROOT")
        result_dir    = os.path.join(prj_root,'dv/tb/reg_model/sv')
        return result_dir

def get_output_ral_dir():
    if sys.platform == 'linux':
        prj_root = os.getenv("PRJ_ROOT")
        result_dir    = os.path.join(prj_root,'dv/tb/reg_model/ral')
        return result_dir


class St_CPU:
    def __init__(self, name):
        self.name = name
        self.revision = ''
        self.derivedFrom = ''
        self.endian = 'little'
        self.srsPresent = False
        self.mpuPresent = False
        self.fpuPresent = False
        self.dspPresent = False
        self.icachePresent = False
        self.dcachePresent = False
        self.mmu = False
        self.itcmPresent = False
        self.dtcmPresent = False
        self.l2cachePresent = False

    def get_inst_str(self):
        return f'name:{self.name},derivedFrom:{self.derivedFrom},revision:{self.revision},endian:{self.endian},srs:{self.srsPresent},mpu:{self.mpuPresent},fpu:{self.fpuPresent},dsp:{self.dspPresent},icache:{self.icachePresent},dcache:{self.dcachePresent},itcm:{self.itcmPresent},dtcm:{self.dtcmPresent},l2cache:{self.l2cachePresent}'

    def toJson(self):
        json_str = '{\n'+f'"name": "{self.name}","revision": "{self.revision}","derivedFrom": "{self.derivedFrom}","endian": "{self.endian}","srsPresent": {getBoolStr(self.srsPresent)},"mpuPresent": {getBoolStr(self.mpuPresent)},"fpuPresent": {getBoolStr(self.fpuPresent)}'
        json_str += f',"dspPresent": {getBoolStr(self.dspPresent)},"icachePresent": {getBoolStr(self.icachePresent)},"dcachePresent": {getBoolStr(self.dcachePresent)},"mmu": {getBoolStr(self.mmu)},"itcmPresent": {getBoolStr(self.itcmPresent)},"dtcmPresent": {getBoolStr(self.dtcmPresent)},"l2cachePresent": {getBoolStr(self.l2cachePresent)}' + '\n}'
        return json_str

class St_Interrupt:
    def __init__(self, name):
        self.name = name
        self.value = 0
        self.description = ''

    def get_inst_str(self):
        return f'Interrupt:  name:{self.name},value:{self.value},desc:{self.description}'
    
    def toJson(self): 
        return '{\n' + f'"name":"{self.name}","value": {self.value},"description": "{self.description}"' + '\n}'


class St_AddressBlock:
    def __init__(self, offset):
        self.offset = offset
        self.size = 0
        self.usage = ''

    def get_inst_str(self):
        return f'AddressBlock:  offset:{self.offset},size:{self.size},usage:{self.usage}'
    
    def toJson(self): 
        return '{\n' + f'"offset": "{self.offset}","size": "{self.size}","usage": "{self.usage}"' + '\n}'


class St_Memory:
    def __init__(self, name):
        self.name = name
        self.derivedFrom = ''
        self.baseAddress = 0
        self.addressOffset = 0
        self.size = 0
        self.access = ''
        self.usage = ''
        self.processor = ''
        self.description = ''

    def get_inst_str(self):
        return f'Memory:  name:{self.name},derivedFrom:{self.derivedFrom},addrBase:{self.baseAddress},addrOffset:{self.addressOffset},size:{self.size},access:{self.access},usage:{self.usage},proc:{self.processor},desc:{self.description}'

    def toJson(self):
        json_str = '{\n'+f'"name": "{self.name}","derivedFrom": "{self.derivedFrom}","baseAddress": "{self.baseAddress}","addressOffset": "{self.addressOffset}","size": {self.size},"access": "{self.access}","usage": "{self.usage}","processor": "{self.processor}","description": "{self.description}"' + '\n}'
        return json_str

class St_Enum_Val:
    def __init__(self, name:str, val):
        self.name = name
        self.desc = ''
        self.value = val

    def get_inst_str(self):
        return f'enum {self.name} = {self.value}\t // {self.desc}'
    
    def toJson(self):
        json_str = '{\n'+f'"name": "{self.name}","description": "{self.desc}","value": "{self.value}"' + '\n}'
        return json_str


class St_Field:
    def __init__(self, name:str):
        self.name = name
        self.description = ''
        self.bitOffset = 0
        self.bitWidth = 1
        self.access = 'RW'
        self.defaultValue = ''
        self.enumValues = []
        self.writeConstraint = ''  # enum or range
        self.range_min = None
        self.range_max = None
        self.hdl_path = ''

    def get_inst_str(self):
        inst_str = 'Field:\n'
        inst_str += f'\tname:{self.name},desc:{self.description},offset:{self.bitOffset},bitWid:{self.bitWidth},access:{
            self.access},defaultVal:{self.defaultValue},writeConst:{self.writeConstraint},range:[{self.range_min}:{self.range_max}]'
        if self.enumValues:
            inst_str += '\n\tenum: \n'
            for e in self.enumValues:
                inst_str += '\t'+e.get_inst_str()+'\n'
        return inst_str
    
    def toJson(self) -> str:
        json_str = '{\n' f'"name": "{self.name}","description": "{self.description}","bitOffset": {self.bitOffset}, "bitWidth": {self.bitWidth},"access": "{self.access}","defaultValue": "{self.defaultValue}"'
        json_str += f',"writeConstraint": "{self.writeConstraint}","hdl_path": "{self.hdl_path}"'
        if self.writeConstraint == 'enumerated':
            if self.enumValues:
                json_str += ',"enumValues": [\n'
                bFirstEnum = True
                for e in self.enumValues:
                    if not bFirstEnum:
                        json_str += ','
                        pass
                    json_str += e.toJson() + '\n'
                    bFirstEnum = False
                    pass
                json_str += '\n]'
                pass
            pass
        elif self.writeConstraint == 'range':
            if self.range_max and self.range_min:
                json_str += ',"range": {\n' + f'"minimum": "{self.range_min}","maximum": "{self.range_max}"' + '\n}'
                pass
            pass
        json_str += '\n}'
        return json_str


class St_Register:
    def __init__(self, name:str, access:str, size = 32):
        self.name = name
        self.dim = 0
        self.dimIncrement = 0
        self.dimName = ''
        self.description = ''
        self.headRegisterName = ''
        self.alternateRegister = ''   # 别名，
        self.alternateGroupName = ''
        self.addressOffset = ''
        self.size = size
        self.access = access
        self.resetValue = ''
        self.resetMask = ''
        self.dataType = 'uint32_t'
        self.fields = []
        self.hdl_path = ''
        self.nValidFdCount = 0

    def getRegName(self) -> str:
        cRegName = self.name
        if self.headRegisterName:
            cRegName = self.headRegisterName
        return cRegName

    def get_inst_str(self) -> str:
        inst_str = 'Register:\n'
        inst_str += f'name:{self.name},dim:{self.dim},dimIncr:{self.dimIncrement},desc:{self.description},cRegNmae:{self.headRegisterName},alterReg:{
            self.alternateRegister},alterGroup:{self.alternateGroupName},offset:{self.addressOffset},size:{self.size},access:{self.access}\n'
        for e in self.fields:
            inst_str += '\t'+e.get_inst_str()+'\n'
        return inst_str
    
    def toJson(self) -> str:
        json_str = '{\n' f'"name": "{self.name}","dim": {self.dim},"dimIncrement": {self.dimIncrement}, "dimName": "{self.dimName}","description": "{self.description}","headRegisterName": "{self.headRegisterName}"'
        json_str += f',"alternateRegister": "{self.alternateRegister}","alternateGroupName": "{self.alternateGroupName}","addressOffset": "{self.addressOffset}","size": {self.size},"access": "{self.access}"'
        json_str += f',"resetValue": "{self.resetValue}","resetMask": "{self.resetMask}","hdl_path": "{self.hdl_path}"'
        if self.fields:
            bFristFd = True
            json_str += ',"fields": [\n'
            for f in self.fields:
                if f.name.upper() == 'RESERVED': #Reserved
                    continue
                if not bFristFd:
                    json_str += ','
                    pass
                json_str += f.toJson() + '\n'
                bFristFd = False
                pass
            json_str += '\n]'
            pass
        json_str += '\n}'
        return json_str


class St_Cluster:
    def __init__(self, name:str):
        self.name = name
        self.dim = 0
        self.dimIncrement = 0
        self.dimName = ''
        self.description = ''
        self.addressOffset = ''
        self.alternateCluster = ''   # 别名，
        self.alternateGroupName = ''
        self.headerStructName = ''
        self.clusters = []  # 子cluster或者register
        self.rowStart = 0
        self.rowEnd = 0

    def get_inst_str(self):
        inst_str = 'Cluster:\n'
        inst_str += f'name:{self.name},dim:{self.dim},dimIncr:{self.dimIncrement},desc:{self.description},addrOffser:{
            self.addressOffset},alter:{self.alternateCluster},alterGroupName:{self.alternateGroupName},cStructName:{self.headerStructName}\n'
        for e in self.clusters:
            inst_str += '\t'+e.get_inst_str()+'\n'
        return inst_str
    
    def toJson(self):
        json_str = '{\n' f'"name": "{self.name}","dim": {self.dim},"dimIncrement": {self.dimIncrement}, "dimName": "{self.dimName}","description": "{self.description}","headerStructName": "{self.headerStructName}"'
        json_str += f',"alternateCluster": "{self.alternateCluster}","alternateGroupName": "{self.alternateGroupName}"'
        if self.clusters:
            clu_lst = []
            reg_lst = []
            for item in self.clusters:
                if isinstance(item,St_Cluster):
                    clu_lst.append(item)
                    pass
                elif isinstance(item,St_Register):
                    reg_lst.append(item)
                    pass
            if clu_lst:
                bFristClu = True
                json_str += ',"clusters": [\n'
                for c in clu_lst:
                    if not bFristClu:
                        json_str += ','
                        pass
                    json_str += c.toJson() + '\n'
                    bFristClu = False
                    pass
                json_str += '\n]'
                pass
            if reg_lst:
                bFristReg = True
                json_str += ',"registers": [\n'
                for r in reg_lst:
                    if r.name.upper() == 'RESERVED':
                        continue
                    if not bFristReg:
                        json_str += ','
                        pass
                    json_str += r.toJson() + '\n'
                    bFristReg = False
                    pass
                json_str += '\n]'
                pass
            pass
        json_str += '\n}'
        return json_str


class St_Peripheral:
    def __init__(self, name):
        self.name = name
        self.derivedFrom = ''
        self.processor = ''
        self.aliasPeripheral = ''
        self.prefixToName = ''
        self.suffixToName = ''
        self.moduleName = ''
        self.instanceName = ''
        self.description = ''
        self.busInterface = ''
        self.headerStructName = ''
        self.baseAddress = ''
        self.addrDerivedFrom = ''
        self.addressOffset = ''
        self.addrBlocksRowindex = 0
        self.addressBlocks = []
        self.interuptsRowindex = 0
        self.interrupts = []
        self.clust_reg_lst = []  # cluster or register

    def get_inst_str(self):
        inst_str = 'Peripheral: \n'
        inst_str += f'name:{self.name}, derivedFrom:{self.derivedFrom},proc:{self.processor},alias:{self.aliasPeripheral},moduleName:{self.moduleName},instName:{self.instanceName},hStructName:{
            self.headerStructName},prefix:{self.prefixToName},suffix:{self.suffixToName},addrBlocks:{self.addrBlocksRowindex},interupt:{self.interuptsRowindex},desc:{self.description}\n'
        for adb in self.addressBlocks:
            inst_str += '\t'+adb.get_inst_str()+'\n'
        for interu in self.interrupts:
            inst_str += '\t'+interu.get_inst_str()+'\n'
        return inst_str
    
    def getAddrStr(self):
        addr_str = self.baseAddress
        if self.addrDerivedFrom:
            addr_str = self.addrDerivedFrom 
        addr_str += ' + '+ self.addressOffset
        return addr_str
    
    def getDirectAddrStr(self):
        addr_str = self.baseAddress
        addr_str += ' + '+ self.addressOffset
        return addr_str


    
    def toJson(self):
        json_str = '{\n' f'"name": "{self.name}","derivedFrom": "{self.derivedFrom}","processor": "{self.processor}", "aliasPeripheral": "{self.aliasPeripheral}","prefixToName": "{self.prefixToName}","suffixToName": "{self.suffixToName}"'
        json_str += f',"moduleName": "{self.moduleName}","instanceName": "{self.instanceName}","description": "{self.description}","busInterface": "{self.busInterface}","headerStructName": "{self.headerStructName}"'
        json_str += f',"baseAddress": "{self.baseAddress}","addrDerivedFrom": "{self.addrDerivedFrom}","addressOffset": "{self.addressOffset}"'
        if self.addressBlocks:
            json_str += ',"addressBlocks": [\n'
            bNotFirst = False
            for a in self.addressBlocks:
                if bNotFirst:
                    json_str += ','
                json_str += a.toJson()
                bNotFirst = True
                pass
            json_str += '\n]'
        if self.interrupts:
            json_str += ',"interupts": [\n'
            bNotFirst = False
            for i in self.interrupts:
                if bNotFirst:
                    json_str += ','
                json_str += i.toJson()
                bNotFirst = True
                pass
            json_str += '\n]'
        if not self.derivedFrom:
            # 不需要另外生成寄存器等信息
            if self.clust_reg_lst:
                clu_lst = []
                reg_lst = []
                for item in self.clust_reg_lst:
                    if isinstance(item,St_Cluster):
                        clu_lst.append(item)
                        pass
                    elif isinstance(item,St_Register):
                        reg_lst.append(item)
                        pass
                if clu_lst:
                    bFristClu = True
                    json_str += ',"clusters": [\n'
                    for c in clu_lst:
                        if not bFristClu:
                            json_str += ','
                            pass
                        json_str += c.toJson() + '\n'
                        bFristClu = False
                        pass
                    json_str += '\n]'
                    pass
                if reg_lst:
                    bFristReg = True
                    json_str += ',"registers": [\n'
                    for r in reg_lst:
                        if r.name.upper() == 'RESERVED':
                            continue
                        if not bFristReg:
                            json_str += ','
                            pass
                        json_str += r.toJson() + '\n'
                        bFristReg = False
                        pass
                    json_str += '\n]'
                    pass
                pass
        json_str += '\n}'
        return json_str


class St_Device:
    def __init__(self, name):
        self.name = name
        self.vendor = ''
        self.version = ''
        self.series = ''
        self.description = ''
        self.cpus = []
        self.width = 32
        self.memories = []
        self.peripherals = {}
        self.interrupts = set()

    def get_inst_str(self):
        inst_str = 'Device: \n'
        inst_str += f'name:{self.name}, vendor:{self.vendor}, version:{
            self.version}, series:{self.series}, desc: {self.description}\n'
        for cpu in self.cpus:
            inst_str += '\t'+cpu.get_inst_str()+'\n'
        for mem in self.memories:
            inst_str += '\t'+mem.get_inst_str()+'\n'
        return inst_str
    def toJson(self):
        json_str = '{\n' f'"name": "{self.name}","vendor": "{self.vendor}","version": "{self.version}", "series": "{self.series}","description": "{self.description}","width": {self.width}'
        if self.cpus:
            json_str += ',"cpus": [\n'
            bNotFirst = False
            for c in self.cpus:
                if bNotFirst:
                    json_str += ','
                json_str += c.toJson()
                bNotFirst = True
                pass
            json_str += '\n]'
        if self.memories:
            json_str += ',"memories": [\n'
            bNotFirst = False
            for m in self.memories:
                if bNotFirst:
                    json_str += ','
                json_str += m.toJson()
                bNotFirst = True
                pass
            json_str += '\n]'
        # if self.interrupts:
        #     json_str += ',"interupts": [\n'
        #     bNotFirst = False
        #     for i in self.interrupts:
        #         if bNotFirst:
        #             json_str += ','
        #         json_str += i.toJson()
        #         bNotFirst = True
        #         pass
        #     json_str += '\n]'
        if self.peripherals:
            json_str += ',"peripherals": [\n'
            bNotFirst = False
            for p in self.peripherals:
                plst = self.peripherals[p]
                if plst:
                    for p_p in plst:
                        if bNotFirst:
                            json_str += ','
                        json_str += p_p.toJson()
                        bNotFirst = True
                pass
            json_str += '\n]'
        json_str += '\n}'
        return json_str


class St_ClusterInnerRange:
    def __init__(self, start) -> None:
        self.rowStart = start
        self.rowEnd = 0
        self.parentClustRow = 0

    def get_inst_str(self):
        return f'rowStart: {self.rowStart}, rowEnd: {self.rowEnd}, parentClusetrRow:{self.parentClustRow}'


class St_clusterFlag_info:
    def __init__(self) -> None:
        self.row = 0
        self.bEnd = False


def markCell_InvalidFunc2(ws:worksheet, row:int, col:int, clr:str ='ff0000'):
    double = Side(border_style="double", color=clr)
    border = Border(left=double,
                    right=double,
                    top=double,
                    bottom=double)
    cell = ws.cell(row, col)
    cell.border = border
    cell.font = Font(color="FF0000")


def markCell_InvalidFunc(ws:worksheet, cellstr:str, clr: str ='ff0000'):
    double = Side(border_style="double", color=clr)
    border = Border(left=double,
                    right=double,
                    top=double,
                    bottom=double)
    cell = ws[cellstr]
    cell.border = border
    cell.font = Font(color="FF0000")


def isHexString(strVal:str, b0xStart:bool = True):
    strUpper = strVal.upper()
    brt = True
    strhex = strUpper
    if b0xStart:
        brt = strUpper.startswith('0X')
        strhex = strUpper[2:]
    hexstr = '0123456789ABCDEF'
    if brt:
        for c in strhex:
            if c not in hexstr:
                brt = False
                break
    return brt

def getIntValFromHexString(strVal:str,b0xStart:bool = True):
    nPos = 0
    if b0xStart:
        nPos =2
    nVal = int(strVal[nPos:],16)
    return nVal

def isUnallowedVarName(strVal:str):
    # strVal = strVal.strip()
    pattern = '^[a-zA-Z_][a-zA-Z0-9_]*$'
    matchObject = re.search(pattern, strVal)
    # if matchObject is None:
    #     print('%s is not Id' % id)
    # else:
    #     print('%s is Id' % id)
    return (matchObject is None)


def checkModuleSheetValue(ws:worksheet, sheetName:str):  # 传入worksheet
    print("ModuleSheet : "+sheetName)
    bMod_CheckErr = False
    perip_title = ws['A1'].value
    if perip_title != 'peripheral:':
        markCell_InvalidFunc(ws, 'A1')
        print('peripheral: must be in the A1 cell!')
        return False
    # 从第三行开始读取直到空行
    nRows = ws.max_row
    maxCols = ws.max_column

    flag_names = ('addressBlocks:', 'interrupts:',
                  'cluster:', 'end cluster', 'register:')
    flag_dict = {}
    flag_rows = []
    emptyA_rows = []
    perip_list = []
    interupts_set = set()
    row = 3

    while row <= nRows:
        val = ws.cell(row, 1).value
        if val in flag_names:
            flag_rows.append(row)
            if val in flag_dict:
                flag_dict[val].append(row)
            else:
                flag_dict[val] = [row]
            pass
        elif not val:
            bEmpty = True
            for col in range(2,6):
                val = ws.cell(row, col).value
                if val:
                    bEmpty = False
                    break
            if bEmpty:
                emptyA_rows.append(row)
        row += 1
    if flag_dict:
        print(flag_dict)
    # flag_rows.sort()
    row = 3
    row_end = nRows
    if flag_rows:
        row_end = flag_rows[0]
    index = 0
    mod_name = ''
    struct_name = ''
    prefix_str = ''
    suffix_str = ''
    while row < row_end:
        if row not in emptyA_rows:
            name = ws.cell(row, 2).value
            st_perip = St_Peripheral(name)
            alias = ws.cell(row, 7).value
            derivedFrom = ws.cell(row, 1).value
            if alias:
                for perip in perip_list:
                    if perip.name == alias:
                        st_perip = copy.copy(perip)
                        break
            else:
                if derivedFrom:
                    for perip in perip_list:
                        if perip.name == derivedFrom:
                            st_perip = copy.copy(perip)
                            break
            if alias:
                st_perip.aliasPeripheral = alias
            if derivedFrom:
                st_perip.derivedFrom = derivedFrom
            st_perip.name = name

            addrDFrom = ws.cell(row, 3).value
            if addrDFrom:
                st_perip.addrDerivedFrom = addrDFrom
            baseAddr = ws.cell(row, 4).value
            if baseAddr:
                st_perip.baseAddress = baseAddr
            offset = ws.cell(row,5).value
            if offset:
                st_perip.addressOffset = offset
            processor = ws.cell(row, 6).value
            if processor:
                st_perip.processor = processor
            mod_inst_name = ws.cell(row, 8).value
            if index == 0 and mod_inst_name:
                mod_name = mod_inst_name
            if mod_inst_name and mod_inst_name != mod_name:
                print(f'moduleName must be same at row: {row}')
                markCell_InvalidFunc2(ws, row, 8)
                bMod_CheckErr = True
            if mod_inst_name:
                st_perip.moduleName = mod_name
            inst_Name = ws.cell(row, 9).value
            if inst_Name:
                st_perip.instanceName = inst_Name
            busInf = ws.cell(row,10).value
            if busInf:
                st_perip.busInterface = busInf

            mod_inst_structName = ws.cell(row, 11).value
            if index == 0 and mod_inst_structName:
                struct_name = mod_inst_structName
            if mod_inst_structName and mod_inst_structName != struct_name:
                print(f'headerStructName must be same at row: {row}')
                markCell_InvalidFunc2(ws, row, 8)
                bMod_CheckErr = True
            if mod_inst_structName:
                st_perip.headerStructName = struct_name

            mod_inst_prefix = ws.cell(row, 12).value
            if index == 0 and mod_inst_prefix:
                prefix_str = mod_inst_prefix
            if mod_inst_prefix and mod_inst_prefix != prefix_str:
                print(f'prefixToName must be same at row: {row}')
                markCell_InvalidFunc2(ws, row, 12)
                bMod_CheckErr = True
            if prefix_str:
                st_perip.prefixToName = prefix_str

            mod_inst_suffix = ws.cell(row, 13).value
            if index == 0 and mod_inst_suffix:
                suffix_str = mod_inst_suffix
            if mod_inst_suffix and mod_inst_suffix != suffix_str:
                print(f'suffixToName must be same at row: {row}')
                markCell_InvalidFunc2(ws, row, 13)
                bMod_CheckErr = True
            if suffix_str:
                st_perip.suffixToName = suffix_str
            desc = ws.cell(row, 16).value
            if desc:
                st_perip.description = desc
            addri = ws.cell(row, 14).value
            if addri:
                bIn = False
                if isinstance(addri, str) and addri.startswith('A'):
                    r_i = addri[1:]
                    if r_i.isdecimal():
                        st_perip.addrBlocksRowindex = int(r_i)
                        if st_perip.addrBlocksRowindex in flag_dict['addressBlocks:']:
                            bIn = True
                if not bIn:
                    print(
                        f'addressBlocks must be the cell of addressBlocks at row: {row}')
                    markCell_InvalidFunc2(ws, row, 14)
                    bMod_CheckErr = True

            interi = ws.cell(row, 15).value
            if interi:
                bIn = False
                if isinstance(interi, str) and interi.startswith('A'):
                    r_i = interi[1:]
                    if r_i.isdecimal():
                        st_perip.interuptsRowindex = int(r_i)
                        if st_perip.interuptsRowindex in flag_dict['interrupts:']:
                            bIn = True
                if not bIn:
                    print(
                        f'interrupts must be the cell of interrupts at row: {row}')
                    markCell_InvalidFunc2(ws, row, 15)
                    bMod_CheckErr = True
            perip_list.append(st_perip)
        row += 1

    # 处理各Flag
    flags_count = len(flag_rows)
    addrblocks_flag_lst = flag_dict['addressBlocks:']
    addrb_dict = {}
    if addrblocks_flag_lst:
        for a_i in addrblocks_flag_lst:
            addrb_lst = []
            row_end = nRows
            f_i = flag_rows.index(a_i)
            if f_i < flags_count-1:
                row_end = flag_rows[f_i+1] - 1
            # 读取addrblocks
            #先读出CPU的各列对应的字段
            row_addr_header =a_i+1
            addrb_col_dict = {}
            for col_i in range(1,maxCols):
                val = ws.cell(row_addr_header,col_i).value
                if val:
                    addrb_col_dict[val] = col_i
            # print(addrb_col_dict)
            row = a_i+2
            while row <= row_end:
                if row not in emptyA_rows:
                    offset = ws.cell(row, addrb_col_dict['offset']).value
                    size = ws.cell(row, addrb_col_dict['size']).value
                    usage = ws.cell(row, addrb_col_dict['usage']).value
                    st_addrb = St_AddressBlock(offset)
                    st_addrb.size = size
                    st_addrb.usage = usage
                    addrb_lst.append(st_addrb)
                row += 1
            addrb_dict[a_i] = addrb_lst

    interrupts_flag_lst = flag_dict['interrupts:']
    interu_dict = {}
    if interrupts_flag_lst:
        for a_i in interrupts_flag_lst:
            interu_lst = []
            row_end = nRows
            f_i = flag_rows.index(a_i)
            if f_i < flags_count-1:
                row_end = flag_rows[f_i+1] - 1
            # 读取interrupts
            #先读出CPU的各列对应的字段
            row_itq_header =a_i+1
            itq_col_dict = {}
            for col_i in range(1,maxCols):
                val = ws.cell(row_itq_header,col_i).value
                if val:
                    itq_col_dict[val] = col_i
            # print(itq_col_dict)
            row = a_i+2
            while row <= row_end:
                if row not in emptyA_rows:
                    name = ws.cell(row, itq_col_dict['name']).value
                    val = ws.cell(row, itq_col_dict['number']).value
                    desc = ws.cell(row, itq_col_dict['description']).value
                    st_inter = St_Interrupt(name)
                    st_inter.value = val
                    st_inter.description = desc
                    interu_lst.append(st_inter)
                    interupts_set.add(st_inter)
                row += 1
            interu_dict[a_i] = interu_lst

    if perip_list:
        for p in perip_list:
            if isinstance(p, St_Peripheral):
                if p.addrBlocksRowindex in addrb_dict:
                    p.addressBlocks = addrb_dict[p.addrBlocksRowindex]
                if p.interuptsRowindex in interu_dict:
                    p.interrupts = interu_dict[p.interuptsRowindex]
                # print(p.get_inst_str())

    cluster_lst = flag_dict['cluster:']
    end_cluster_lst = flag_dict['end cluster']
    st_clu_reg_list = []
    # parent_clu_reg_list = None
    # parent_clu_ = None
    if isinstance(end_cluster_lst, list) and len(cluster_lst) != len(end_cluster_lst):
        bMod_CheckErr = True
        print("Error:  cluster not all have end cluster")
    else:
        cluster_len = len(cluster_lst)
        # 将cluseter和endcluster按行号循序重新排列
        clu_item_lst = []
        for c in cluster_lst:
            clu_ = St_clusterFlag_info()
            clu_.row = c
            clu_item_lst.append(clu_)
        for e in end_cluster_lst:
            clu_ = St_clusterFlag_info()
            clu_.row = e
            clu_.bEnd = True
            bInsert = False
            for index in range(len(clu_item_lst)):
                if clu_item_lst[index].row > e:
                    clu_item_lst.insert(index, clu_)
                    bInsert = True
                    break
            if not bInsert:
                clu_item_lst.append(clu_)
        # 现在得到了顺序排列
        clu_range_list = []
        clu_item_len = len(clu_item_lst)
        cur_clu_range = None
        parent_clu_range_lst = []
        for index in range(clu_item_len):
            clu_item = clu_item_lst[index]
            if cur_clu_range:
                if clu_item.bEnd:
                    if cur_clu_range:
                        cur_clu_range.rowEnd = clu_item.row
                        if parent_clu_range_lst:
                            cur_clu_range = parent_clu_range_lst.pop()
                        else:
                            cur_clu_range = None
                else:
                    if cur_clu_range.rowEnd == 0:
                        clu_range = St_ClusterInnerRange(clu_item.row)
                        clu_range.parentClustRow = cur_clu_range.rowStart
                        parent_clu_range_lst.append(cur_clu_range)
                        cur_clu_range = clu_range
                        clu_range_list.append(clu_range)
            else:
                cur_clu_range = clu_range = St_ClusterInnerRange(clu_item.row)
                clu_range_list.append(clu_range)

        for clu_range in clu_range_list:
            print(clu_range.get_inst_str())
            if clu_range.parentClustRow == 0:
                st_clu_reg_list, row, bError = readCluster(ws, st_clu_reg_list, clu_range, clu_range_list)
                if bError:
                    bMod_CheckErr = True

    reg_row_lst = flag_dict['register:']
    if reg_row_lst:
        for a_i in reg_row_lst:
            # 先判断是否已经在cluster内
            bInCluster = False
            for clu_range in clu_range_list:
                if clu_range.parentClustRow == 0 and a_i in range(clu_range.rowStart, clu_range.rowEnd):
                    bInCluster = True
                    break
            if bInCluster:
                continue
            row_end = nRows
            f_i = flag_rows.index(a_i)
            if f_i < flags_count-1:
                row_end = flag_rows[f_i+1] - 1
            st_clu_reg_list, row ,bError = readRegister(ws,  row_end, a_i, st_clu_reg_list)
            if bError:
                bMod_CheckErr = True


    if bMod_CheckErr:
        print("ModuleSheet : "+sheetName + ' have Errors')
    else:
        print("ModuleSheet : "+sheetName + '  Ok')
        # for e in st_clu_reg_list:
        #     print(e.get_inst_str())
        for p in perip_list:
            p.clust_reg_lst = st_clu_reg_list

    return not bMod_CheckErr, perip_list,interupts_set


def readRegister(ws:worksheet, row_end:int, row_start:int, parent_clu_reg_list:list):
    bError = False
    maxCols = ws.max_column
    #先读出CPU的各列对应的字段
    row_reg_header = row_start + 1
    reg_col_dict = {}
    for col_i in range(1,maxCols):
        val = ws.cell(row_reg_header,col_i).value
        if val:
            reg_col_dict[val] = col_i
    # print(reg_col_dict)
    regFd_col_dict = {}
    row = row_start+2
    bField = False
    cur_st_reg = None
    cur_reg_fd = None
    flag_names = ('cluster:', 'end cluster', 'register:')
    while row <= row_end:
        value = ws.cell(row, 1).value
        if value:
            if value in flag_names:
                break
            if not bField:
                name = value
                if name == 'fields:':  # 开始处理 field
                    bField = True
                    row += 1
                    for col_i in range(1,maxCols):
                        val = ws.cell(row,col_i).value
                        if val:
                            regFd_col_dict[val] = col_i
                    # print(regFd_col_dict)
                else:
                    offset = ws.cell(row, reg_col_dict['addressOffset']).value
                    r_size = ws.cell(row, reg_col_dict['size']).value
                    if r_size:
                        if isinstance(r_size,str):
                            r_size = int(r_size)
                    size = 32
                    if isinstance(r_size,int):
                        size = r_size
                    access = ws.cell(row, reg_col_dict['access']).value
                    cur_st_reg = St_Register(name, access, size)
                    cur_st_reg.addressOffset = offset
                    #addrOffset = int(offset[2:], 16)
                    addrOffset = int(offset, 16)
                    resetValue = ws.cell(row, reg_col_dict['resetValue']).value
                    if resetValue:
                        cur_st_reg.resetValue = resetValue
                    resetMask = ws.cell(row, reg_col_dict['resetMask']).value
                    if resetMask:
                        cur_st_reg.resetMask = resetMask
                    headRegisterName = ws.cell(row, reg_col_dict['headRegisterName']).value
                    if headRegisterName:
                        cur_st_reg.headRegisterName = headRegisterName
                    alternateRegister = ws.cell(row, reg_col_dict['alternateRegister']).value
                    if alternateRegister:
                        cur_st_reg.alternateRegister = alternateRegister
                    alternateGroupName = ws.cell(row, reg_col_dict['alternateGroupName']).value
                    if alternateGroupName:
                        cur_st_reg.alternateGroupName = alternateGroupName
                    dim = ws.cell(row, reg_col_dict['dim']).value
                    if dim: 
                        cur_st_reg.dim = dim
                    dimIncrement = ws.cell(row, reg_col_dict['dimIncrement']).value
                    if isinstance(dimIncrement,str):
                        dimIncrement = int (dimIncrement)
                    if isinstance(dimIncrement,int):
                        cur_st_reg.dimIncrement = dimIncrement
                    dimName = ws.cell(row, reg_col_dict['dimName']).value
                    if dimName:
                        cur_st_reg.dimName = dimName
                    desc = ws.cell(row, reg_col_dict['description']).value
                    if desc:
                        cur_st_reg.description = desc
                    hdl_path = ws.cell(row, reg_col_dict['pathHDL']).value
                    if hdl_path:
                        cur_st_reg.hdl_path = hdl_path
                    if parent_clu_reg_list:
                        i = 0
                        for clu_reg in parent_clu_reg_list:
                            offset = int(clu_reg.addressOffset[2:], 16)
                            if offset > addrOffset:
                                parent_clu_reg_list.insert(i, cur_st_reg)
                                break
                            i += 1
                        if i == len(parent_clu_reg_list):
                            parent_clu_reg_list.append(cur_st_reg)
                    else:
                        parent_clu_reg_list.append(cur_st_reg)
            else:
                name = ws.cell(row, regFd_col_dict['bitName']).value
                if not name:
                    row += 1
                    continue
                bSameFieldAsPrev = False
                if name and name.upper() != 'RESERVED':
                    for f in cur_st_reg.fields:
                        if f.name == name:
                            if f == cur_reg_fd:
                                bSameFieldAsPrev = True
                            else:
                                print(f'In Register Field Name not allow repeat at Row {row}')
                                bError = True
                            break
                if not bSameFieldAsPrev:
                    cur_reg_fd = fd = St_Field(name.upper())
                    bitrange = ws.cell(row, regFd_col_dict['bitRange']).value
                    end, op, start = bitrange.partition(':')
                    start_bit = start[0:-1]
                    end_bit = end[1:]
                    offset = nStart_bit = int(start_bit)
                    nEnd_bit = int(end_bit)
                    bitWidth = nEnd_bit - nStart_bit+1
                    access = ws.cell(row, regFd_col_dict['access']).value
                    defaultVal = ws.cell(row, regFd_col_dict['defaultValue']).value
                    writeConstraint = ws.cell(row, regFd_col_dict['writeConstraint']).value
                    range_min = ws.cell(row, regFd_col_dict['minimum']).value
                    range_max = ws.cell(row, regFd_col_dict['maximum']).value
                    enumName = ws.cell(row, regFd_col_dict['enumName']).value
                    enumVal = ws.cell(row, regFd_col_dict['enumValue']).value
                    enumDesc = ws.cell(row, regFd_col_dict['enumDescription']).value
                    desc = ws.cell(row, regFd_col_dict['description']).value
                    hdl_path =ws.cell(row,regFd_col_dict['pathHDL']).value
                    fd.bitOffset = offset
                    fd.bitWidth = bitWidth
                    fd.access = access
                    if defaultVal:
                        fd.defaultValue = defaultVal
                    if writeConstraint:
                        fd.writeConstraint = writeConstraint
                    if range_min:
                        fd.range_min = range_min
                    if range_max:
                        fd.range_max = range_max
                    if hdl_path:
                        fd.hdl_path = hdl_path
                    if desc:
                        fd.description = desc
                    if enumName and enumVal:
                        enum_item = St_Enum_Val(enumName, enumVal)
                        if enumDesc:
                            enum_item.desc = enumDesc
                        fd.enumValues.append(enum_item)
                    # 这里可能需要添加按顺序插入动作
                    binsertFd = False
                    if cur_st_reg.fields:
                        fd_len=len(cur_st_reg.fields)
                        for i in range(fd_len):
                            if cur_st_reg.fields[i].bitOffset > fd.bitOffset:
                                cur_st_reg.fields.insert(i,fd)
                                if fd.name != 'RESERVED':
                                    cur_st_reg.nValidFdCount += 1
                                binsertFd = True
                                break
                    if not binsertFd:
                        cur_st_reg.fields.append(fd)
                        if fd.name != 'RESERVED':
                            cur_st_reg.nValidFdCount += 1
                else:
                    enumName = ws.cell(row, regFd_col_dict['enumName']).value
                    enumVal = ws.cell(row, regFd_col_dict['enumValue']).value
                    if isHexString(enumVal):
                        #nEnumVal = int(enumVal[2:],16)
                        nEnumVal = int(enumVal,16)
                        bitMask = bitWidMask_arr[fd.bitWidth-1]
                        #nBitMask = int(bitMask[2:],16)
                        nBitMask = int(bitMask,16)
                        if nEnumVal >  nBitMask:
                            print(f'Error: In Register Field enum value extends the bitRange at  Row {row}')
                            bError = True
                            pass
                    enumDesc = ws.cell(row, regFd_col_dict['enumDescription']).value
                    if enumName and enumVal:
                        enum_item = St_Enum_Val(enumName, enumVal)
                        if enumDesc:
                            enum_item.desc = enumDesc
                        cur_reg_fd.enumValues.append(enum_item)

        row += 1
    return parent_clu_reg_list, row, bError


def readCluster(ws: worksheet, parent_clu_reg_list:list, clu_range: St_ClusterInnerRange, clu_range_list:list):
    bError = False
    clu_start = clu_range.rowStart
    clu_end = clu_range.rowEnd
    maxCols = ws.max_column
    #先读出CPU的各列对应的字段
    row_cluster_header = clu_start + 1
    cluster_col_dict = {}
    for col_i in range(1,maxCols):
        val = ws.cell(row_cluster_header,col_i).value
        if val:
            cluster_col_dict[val] = col_i
    # print(cluster_col_dict)
    row = clu_start+2
    st_clu = None
    while row < clu_end:
        name = ws.cell(row, 1).value
        if not name:
            row += 1
            continue
        if name == 'cluster:':
            if st_clu:
                for c_r in clu_range_list:
                    if c_r.rowStart == row:
                        st_clu.clusters, row, bError = readCluster(
                            ws, st_clu.clusters, c_r, clu_range_list)
                        break
            else:
                row += 2
            continue
        elif name == 'register:':
            if st_clu:
                st_clu.clusters, row ,bError= readRegister(
                    ws, clu_end, row, st_clu.clusters)
            else:
                row += 2
            continue
        elif name == 'end cluster':
            row += 1
            break
        else:
            addressOffset = ws.cell(row, cluster_col_dict['addressOffset']).value
            if addressOffset:
                st_clu = St_Cluster(name)
                st_clu.rowStart = clu_start
                st_clu.rowEnd = clu_end
                st_clu.addressOffset = addressOffset
                #addrOffset = int(st_clu.addressOffset[2:], 16)
                addrOffset = int(st_clu.addressOffset, 16)
                alternateCluster = ws.cell(row, cluster_col_dict['alternateCluster']).value
                if alternateCluster:
                    st_clu.alternateCluster = alternateCluster
                alternateGroupName = ws.cell(row, cluster_col_dict['alternateGroupName']).value
                if alternateGroupName:
                    st_clu.alternateGroupName = alternateGroupName
                headerStructName = ws.cell(row, cluster_col_dict['headerStructName']).value
                if headerStructName:
                    st_clu.headerStructName = headerStructName
                dim = ws.cell(row, cluster_col_dict['dim']).value
                if dim:
                    st_clu.dim = dim
                dimIncrement = ws.cell(row, cluster_col_dict['dimIncrement']).value
                if dimIncrement:
                    st_clu.dimIncrement = dimIncrement
                description = ws.cell(row, cluster_col_dict['description']).value
                if description:
                    st_clu.description = description
                if parent_clu_reg_list:
                    i = 0
                    for clu_reg in parent_clu_reg_list:
                        #offset = int(clu_reg.addressOffset[2:], 16)
                        offset = int(clu_reg.addressOffset, 16)
                        if offset > addrOffset:
                            parent_clu_reg_list.insert(i, st_clu)
                            break
                        i += 1
                    if i == len(parent_clu_reg_list):
                        parent_clu_reg_list.append(st_clu)
                else:
                    parent_clu_reg_list.append(st_clu)
                pass
        row += 1
    return parent_clu_reg_list, row, bError


# def output_SV_moduleFile(module_inst, modName):
#     out_svh_module_Name = modName.lower()+'_dut_cfg'
#     out_svh_file_name = './'+out_svh_module_Name+'.svh'
#     with open(out_svh_file_name, 'w+') as sv_file:
#         heder_str = f'_{modName.upper()}_DUT_CFG_SVH_'
#         file_str = F'`ifndef {heder_str}\n`define {heder_str}\n\n'

#         file_enum_str = ''

#         uvm_field_str = f'\n\t`uvm_object_utils_begin({out_svh_module_Name})\n'
#         uvm_fd_val_def_str = ""
#         val_def_strarr = ["// Autor: Auto generate by sv", "// Version: 0.0.2 X",
#                           "// Description : set module reg field random value", "// Waring: Do NOT Modify it !", "#pragma once"]
#         for str in val_def_strarr:
#             uvm_fd_val_def_str += f'\t\t$fdisplay(fd, "{str}" );\n'

#         uvm_fd_val_def_str += '\t\t$fdisplay(fd, "   " );\n\n'
#         file_cls_str = f'class {out_svh_module_Name} extends uvm_object;\n\n'
#         for reg in module_inst.reg_list:
#             for fd in reg.field_list:
#                 if fd.bRandom_Enable and fd.attribute.find('W') != -1:
#                     reg_fd_name = f'{reg.reg_name}___{fd.field_name}'
#                     b_fd_enum = False
#                     nbit_Wid = fd.end_bit-fd.start_bit+1
#                     bit_str = 'bit'
#                     if nbit_Wid > 1:
#                         bit_str = f'bit [{nbit_Wid-1}:0]'
#                     if fd.field_enumstr:
#                         # print(fd.field_enumstr)
#                         b_fd_enum = True
#                         enum_lst = fd.field_enumstr.splitlines()
#                         file_enum_str += f'typedef enum {bit_str}'+' {\n'
#                         b_emFirstitem = True
#                         for em in enum_lst:
#                             # print(em)
#                             em_val = em.replace(',', '')
#                             em_val = em_val.strip()
#                             (em_item_name, str, em_item_value) = em_val.partition('=')
#                             em_item_name = em_item_name.strip()
#                             em_item_value = em_item_value.strip().upper()
#                             if not b_emFirstitem:
#                                 file_enum_str += ',\n'
#                             if len(em_item_value) and em_item_value.startswith('0X'):
#                                 em_item_value_int = int(em_item_value, 16)
#                                 file_enum_str += f'\t{em_item_name} {str} {em_item_value_int}'
#                             else:
#                                 file_enum_str += f'\t{em_item_name} {str} {em_item_value}'
#                             b_emFirstitem = False
#                             # file_str
#                         file_enum_str += '\n} '+f'em_{reg_fd_name};\n\n'

#                     if b_fd_enum:
#                         file_cls_str += f'\trand em_{reg_fd_name} {reg_fd_name};\n'
#                     else:
#                         file_cls_str += f'\trand {bit_str}  {reg_fd_name};\n'

#                     uvm_field_str += f'\t\t`uvm_field_int({reg_fd_name}, UVM_ALL_ON)\n'

#                     fd_name_VAL = f'{reg_fd_name.upper()}_VALUE_'
#                     fd_name_VAL = fd_name_VAL.ljust(48)
#                     if b_fd_enum:
#                         uvm_fd_val_def_str += f'\t\t$fdisplay(fd, "#define \t {fd_name_VAL}   0x%X   //%s",  {reg_fd_name}, {reg_fd_name}.name());\n'
#                     else:
#                         uvm_fd_val_def_str += f'\t\t$fdisplay(fd, "#define \t {fd_name_VAL}   0x%X",  {reg_fd_name});\n'
#         uvm_field_str += f'\t`uvm_object_utils_end\n'

#         uvm_field_str += f'\n\tfunction new(string name = "{out_svh_module_Name}");\n'
#         uvm_field_str += """\t\tsuper.new(name);
#     endfunction:new

#     virtual function void print_cfg_to_file();
#         int fd;
# """
#         uvm_field_str += f'\t\tfd = $fopen("{modName}_dut_cfg.h");\n'

#         uvm_field_str += uvm_fd_val_def_str
#         uvm_field_str += """
#         $fclose(fd);
#     endfunction:print_cfg_to_file
# """
#         file_cls_str += uvm_field_str
#         file_cls_str += """endclass

# `endif
# """
#         file_str += file_enum_str+'\n\n'
#         file_str += file_cls_str
#         sv_file.write(file_str)
#         sv_file.close()

#         return out_svh_file_name


def output_C_dev_InOneFile(dev: St_Device):
    pass

def output_C_devFile(dev: St_Device,module_file_lst:list):
    devName = dev.name.lower()
    dev_Name_=dev.name
    if dev_Name_.endswith('xx'):
        dev_Name_=dev_Name_[0:-2]
        dev_Name_=dev_Name_.upper()+'xx'
    else:
        dev_Name_=dev_Name_.upper()
    out_file_name= './'+devName+'.h'
    print('output_C_devFile: ' + dev.name)
    with open(out_file_name, 'w+') as out_file:
        fileHeader = f'/**\n * @file    {devName}.h\n'
        fileHeader += f' * @author  CIP Application Team\n * @brief   {devName} Peripheral Access Layer Header File.'
        fileHeader += """
 *          This file contains:
 *           - Interrupt vector number and
 *           - Data structures and the address mapping for
 *             all peripherals
 *           - Including peripheral's registers declarations and bits
 *             definition Header File
"""

        # 格式化成2016-03-20 11:45:39形式
        today = date.today()
        fileHeader += f' * @version {dev.version} \n * @date    {today.strftime("%y-%m-%d")}\n'

        fileHeader += """
 *
  ******************************************************************************
 * @copyright
 *
"""
        fileHeader += f' *  <h2><center>&copy; Copyright (c){today.year} CIP United Co.\n'
        fileHeader += """
 * All rights reserved.</center></h2>
 *
 * 
 *
 ******************************************************************************
 */

 /*
 * ****************************************************************************
 * ******** Define to prevent recursive inclusion                  ********
 * ****************************************************************************
 */
#pragma once
"""
        fileHeader += f'#ifndef __{dev_Name_}_H\n'
        fileHeader += f'#define __{dev_Name_}_H\n'
        fileHeader += """
/*
 * ============================================================================
 * If building with a C++ compiler, make all of the definitions in this header
 * have a C binding.
 * ============================================================================
 */
#ifdef __cplusplus
extern "C"
{
#endif

/*
 * ****************************************************************************
 * ******** Includes                                                   ********
 * ****************************************************************************
 */
#include <stdint.h>   //for use uint32_t type

/*
 * ****************************************************************************
 * ******** Exported Types                                             ********
 * ****************************************************************************
 */
#if !defined(__ASSEMBLER__)
/*
 * ============================================================================
 * ==             Interrupt vector number defined                            ==
 * ============================================================================
 */
 
"""

# /**
#  * @brief XY2_MP32DIAPT Interrupt Vector Number Definition
#  */
# typedef enum __IRQ_NUMBER
# {
#     CORE_SW0_IRQn       = 0,    /*!< Core timer interrupt       */
#     CORE_SW1_IRQn       = 1,    /*!< Core software interrupt 1  */
#     CORE_HW0_IRQn       = 2,    /*!< Core software interrupt 2  */

#     NUMBER_INT_VECTORS
# } IRQn_t;

        fileHeader +=f'/**\n *  @brief {dev_Name_} Interrupt Vector Number Definition\n */\n'
       
        irq_str = ''
        irq_num_str = ''
        index = 0 
        irCount = len(dev.interrupts) 
        for ir in dev.interrupts:
            index += 1
            if isinstance(ir,St_Interrupt):
                emIr= cst_tab_str+f'{ir.name}'
                emIr = emIr.ljust(cst_emIRQ_SpaceSize)
                emIr += '= '+ str(ir.value)
                if index != irCount:
                    emIr += ','
                else:
                    emIr += ' '
                emIr += cst_tab_str+f'/*!< {ir.description} */\n'
                irq_str += emIr
                irq_num_str += f'#define {ir.name}_NUM'.ljust(cst_IRQ_SpaceSize)  +f'{ir.value}       /*!< {ir.description}       */\n'
        if irq_str:
            fileHeader += 'typedef enum __IRQ_NUMBER\n{\n'   
            fileHeader += irq_str
            fileHeader +='} IRQn_t;\n'
        fileHeader +="""
/**
 * @brief Interrupt handle function typedef
 */
typedef void (*irqFnHandler)(void *);

/**
 * @}
 */

#endif /* !__ASSEMBLER__ */

/*
 * ****************************************************************************
 * ******** Exported constants                                         ********
 * ****************************************************************************
 */
/** @addtogroup Interrupt_Vector_Number Interrupt vector number
 * @{
 */

"""
        fileHeader += irq_num_str
        fileHeader += """
/**
 * @}
 */

 
/*
 * ============================================================================
 * ==             Device Processor Configuration                             ==
 * ============================================================================
 */
/** @addtogroup Configuration_Section_For_PROC
 * @{
 */
#define MIPS_DUAL_CORE              /*!< Dual core feature */

/**
#  * @brief Configuration of the  Processor and Core Peripherals
#  */
"""
        fileHeader += '#define PROC_UNDEFINED\n'
        procName_lst = []
        for cp in dev.cpus:
            cp_name = cp.name.upper()
            procName =  f'PROC_{cp_name}'
            procName_lst.append(cp_name)

            fileHeader += f'#ifdef {procName}\n'
            
            if cp.revision:
                p_index = cp.revision.find('p')
                if p_index !=-1 and p_index < len(cp.revision)-1:
                    cp_str = f'#define __CORE_REV'.ljust(cst_CORE_SpaceSize)
                    rev_str_0 = cp.revision[1:p_index].rjust(2,'0')
                    rev_str_1 = cp.revision[p_index+1:]
                    rev_str_1 = rev_str_1.rjust(2,'0')
                    cp_str += '0x'+rev_str_0+rev_str_1+"U   /*!< Defines processorre's vision               */\n"
                    fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.srsPresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_SRS'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the Shadow regiseters are present or not  */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.mpuPresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_MPU'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the MPU is present or not       */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.fpuPresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_FPU'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the FPU is present or not       */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.dspPresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_DSP'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the DSP is present or not       */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.icachePresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_L1_ICACHE'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the L1 ICache is present or not */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.dcachePresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_L1_DCACHE'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the L1 DCache is present or not */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.mmu:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_MMU_TLB'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the MMU is present or not       */\n"
            fileHeader += cp_str
            val_str = '0U'.ljust(cst_HEXValue_StringSize)
            if cp.l2cachePresent:
                val_str = '1U'.ljust(cst_HEXValue_StringSize)
            cp_str = f'#define __CORE_HAS_L2_CACHE'.ljust(cst_CORE_SpaceSize) + val_str + "/*!< Defines if the L2 Cache is present or not  */\n"
            fileHeader += cp_str


            fileHeader += """
#undef PROC_UNDEFINED
/* Core Access Layer */
#if !defined(__ASSEMBLER__)
"""
            fileHeader += f'#include "core_{cp.name.lower()}.h"\n#endif /* !__ASSEMBLER__ */\n'

            fileHeader += f'#endif //PROC_{cp.name.upper()}\n'
            pass

        
        fileHeader += """
#ifdef PROC_UNDEFINED
    #error "Please #define Processor"
#endif //PROC_UNDEFINED

/**
 * @}
 */
#if !defined(__ASSEMBLER__)
  /* Device system configuration */

#ifndef PROC_UNDEFINED
"""
  
        fileHeader += f'#include "system_{devName}.h" \n'
        fileHeader += """
#endif //PROC_UNDEFINED

#endif /* !__ASSEMBLER__ */

/*
 * ============================================================================
 * ==             Device Memory Space Definition                             ==
 * ============================================================================
 */
/** @addtogroup Peripheral_Memory_Map_Definition
 * @{
 */
/* ----------------------------------------------------------------------------
 * -- Memory Space Map
 * ------------------------------------------------------------------------- */

"""
        mem_proc_dict = {}
        mem_other_str = ''
        for mem in dev.memories:
            if mem.processor:
                mem_str = f'#define {mem.name}'.ljust(cst_Mem_SpaceSize)+f'({mem.baseAddress})'+cst_tab_str
                if mem.description:
                    mem_str += f'/*!< Base address of :{mem.description} - {mem.access} */\n'
                else:
                    mem_str += '\n'
                bInProcLst = False
                for p in procName_lst:
                    if mem.processor == p:
                        bInProcLst = True
                        if p in mem_proc_dict:
                            proc_mem_str = mem_proc_dict[p]
                            proc_mem_str += mem_str
                            mem_proc_dict[p] = proc_mem_str
                        else:
                            mem_proc_dict[p] = mem_str
                        break
                if not bInProcLst:
                    print(f'Error:  MEM {mem.name} processor {mem.processor} not in CPUS. ')
            else:
                if mem.derivedFrom:
                    mem_str = f'#define {mem.name}'.ljust(cst_Mem_SpaceSize)+f'({mem.derivedFrom} + {mem.addressOffset}U)'+cst_tab_str
                    if mem.description:
                        mem_str += f'/*!< Base address of :{mem.description} - {mem.access} */\n'
                    else:
                        mem_str += '\n'
                else:
                    mem_str = f'#define {mem.name}'.ljust(cst_Mem_SpaceSize)+f'({mem.addrBase})'+cst_tab_str
                    if mem.description:
                        mem_str += f'/*!< Base address of :{mem.description} - {mem.access} */\n'
                    else:
                        mem_str += '\n'
                mem_other_str += mem_str
        
        for p in mem_proc_dict:
            fileHeader += f'#ifdef PROC_{p}\n'
            fileHeader += mem_proc_dict[p]
            fileHeader += f'#endif // PROC_{p}\n'
            pass

        fileHeader += """
/* ----------------------------------------------------------------------------
 * -- Peripherals Space Map
 * ------------------------------------------------------------------------- */
"""
        fileHeader += mem_other_str

        fileHeader += """
/* ----------------------------------------------------------------------------
 * -- Special function register map
 * ------------------------------------------------------------------------- */
"""
        perip_proc_dict = {}
        for p_name in dev.peripherals:
            perip_lst=dev.peripherals[p_name]
            for perip in perip_lst:
                bInProcLst = False
                if perip.processor:
                    perip_str = f'#define {perip.name}_BASE'.ljust(cst_Perip_SpaceSize)+f'({perip.getAddrStr()})\n'
                    for p in procName_lst:
                        if perip.processor == p:
                            bInProcLst = True
                            if p in perip_proc_dict:
                                proc_perip_str = perip_proc_dict[p]
                                proc_perip_str += perip_str
                                perip_proc_dict[p] = proc_perip_str
                            else:
                                perip_proc_dict[p] = perip_str
                            break
                        pass
                    pass
                if not bInProcLst:
                    print(f'Error:  perip {perip.name} processor {perip.processor} not in CPUS. ')
                pass
            pass

        for p in perip_proc_dict:
            fileHeader += f'#ifdef PROC_{p}\n'
            fileHeader += perip_proc_dict[p]
            fileHeader += f'#endif // PROC_{p}\n'
            pass

        fileHeader += """

/**
 * @}
 */

#if !defined(__ASSEMBLER__)
/*
 * ============================================================================
 * ==              Peripheral instance declaration                           ==
 * ============================================================================
 */

/** @addtogroup Peripheral_Instance_Declaration
 * @{
 */

/**
 * @}
 */

/*
 * ============================================================================
 * ==              Peripheral register structures & declaration              ==
 * ============================================================================
 */

/** @addtogroup Peripheral_Registers_Bits_Definition
 * @{
 */
/* Device register head file */
"""
        for f in module_file_lst:
            fileHeader += f'#include "{f}"\n'

        fileHeader += """
/*
 * ****************************************************************************
 * ******** Exported macro                                    ********
 * ****************************************************************************
 */

/*
 * ============================================================================
 * Check peripheral Instances
 * ============================================================================
 */

/*
 * ============================================================================
 * Get peripheral Index
 * ============================================================================
 */

/*
 * ****************************************************************************
 * ******** Exported Function                                          ********
 * ****************************************************************************
 */

#endif /* defined(__ASSEMBLER__) */

#ifdef __cplusplus
}
#endif
"""
        fileHeader += f'\n#endif //  __{dev_Name_}_H\n'
        fileHeader +="""
/*
 * ****************************************************************************
 * End File
 * ****************************************************************************
 */

"""

        out_file.write(fileHeader)
    pass


def get_sequence_sv_reg(clu_reg: St_Register,module_index:str,reg_name:str,tab_str:str):
    bAllRegFdHdlPathEmpty = True
    reg_access_str = ''
    regReset_ignore_str =''
    if clu_reg.name.upper() != 'RESERVED':   #Reserved
        regAccess =clu_reg.access
        if regAccess == 'W':
            if clu_reg.dim > 1:
                # for di in range(reg.dim):
                regReset_ignore_str = tab_str + f'for (int i=0; i< {clu_reg.dim}; i++) begin\n'
                regReset_ignore_str += tab_str + cst_tab_str + 'uvm_resource_db#(bit)::set({"REG::",p_sequencer.u_soc_reg_model.'
                regReset_ignore_str += f'{module_index}.{reg_name}[i]'+'.get_full_name()},"NO_REG_HW_RESET_TEST",1,this);\n'
                regReset_ignore_str += tab_str +'end\n\n'
                # pass
            else:
                regReset_ignore_str += tab_str + 'uvm_resource_db#(bit)::set({"REG::",p_sequencer.u_soc_reg_model.'
                regReset_ignore_str += f'{module_index}.{reg_name}'+'.get_full_name()},"NO_REG_HW_RESET_TEST",1,this);\n'
            pass
        bRegFdHdlEmpty = True
        for fd in clu_reg.fields:
            if fd.hdl_path:
                bAllRegFdHdlPathEmpty = False
                bRegFdHdlEmpty = False
        if bRegFdHdlEmpty:
            hdl_empty_str = f'{tab_str}uvm_resource_db#(bit)'+'::set({"REG::",p_sequencer.u_soc_reg_model.'
            hdl_empty_str += module_index
            if clu_reg.dim > 1:
                # for di in range(reg.dim):
                row_hdl_empty_str = tab_str + f'for (int i=0; i< {clu_reg.dim}; i++) begin\n'
                row_hdl_empty_str += cst_tab_str+ hdl_empty_str + f'.{reg_name}[i]'+'.get_full_name()},"NO_REG_ACCESS_TEST",1,this);\n'
                row_hdl_empty_str += tab_str +'end\n\n'
                reg_access_str += row_hdl_empty_str
            else:
                row_hdl_empty_str = hdl_empty_str + f'.{reg_name}' + '.get_full_name()},"NO_REG_ACCESS_TEST",1,this);\n'
                reg_access_str += row_hdl_empty_str
            pass
    # print(regReset_ignore_str)
    # print(reg_access_str)
    return regReset_ignore_str,reg_access_str,bAllRegFdHdlPathEmpty
    

def get_sequence_sv_clu(clu:St_Cluster,module_index:str, parent_clu_name:str,tab_str:str):
    clu_regReset_ignore_str = ''
    clu_regAccess_ignore_str = ''
    bAllRegFdHdlPathEmpty = True
    for clu_reg in clu.clusters:
        clu_name = f'{clu.name}.{clu_reg.name}'
        if parent_clu_name:
            clu_name = parent_clu_name + f'.{clu.name}.{clu_reg.name}'
        # print(clu_name)
        if isinstance(clu_reg,St_Cluster):
            child_reset_str,childstr_access_str, bCluFdHdlPathEmpty = get_sequence_sv_clu(clu_reg,module_index,clu_name,tab_str)
            clu_regReset_ignore_str += child_reset_str
            clu_regAccess_ignore_str += childstr_access_str
            if not bCluFdHdlPathEmpty:
                bAllRegFdHdlPathEmpty = False
            pass
        elif isinstance(clu_reg,St_Register):
            if clu_reg.name.upper() == 'RESERVED':   #Reserved
                continue
            
            regReset_ignore_str,reg_access_str,bRegFdHdlPathEmpty = get_sequence_sv_reg(clu_reg,module_index,clu_name,tab_str)
            if regReset_ignore_str:
                clu_regReset_ignore_str += regReset_ignore_str
            if reg_access_str:
                clu_regAccess_ignore_str += reg_access_str
            if not bRegFdHdlPathEmpty:
                bAllRegFdHdlPathEmpty = False
            pass
        pass
    return clu_regReset_ignore_str, clu_regAccess_ignore_str, bAllRegFdHdlPathEmpty

def output_SequenceSv_moduleFile(preip_lst:list,modName:str,version:str):
    modName = modName.lower()
    out_sv_module_Name = f'{modName}_v_reg_test_sequence'
    out_file_name = out_sv_module_Name+'.sv'
    out_file_Pathname = './uvm/'+out_file_name
    if sys.platform == 'linux':
        out_file_Pathname = os.path.join(get_output_dut_cfg_dir(), out_file_name)
    if preip_lst:
        
        with open(out_file_Pathname, 'w+') as sv_file:
            fileHeader = f'/**\n * @file    {out_file_name}\n'
            fileHeader += f' * @author  CIP Application Team\n # @brief   {modName} sequence UVM test .\n'

            # 格式化成2016-03-20 11:45:39形式
            today = date.today()
            fileHeader += f' # @version {version} \n # @date    {today.strftime("%y-%m-%d")}\n'
            fileHeader += """
*
******************************************************************************
* @copyright
*
"""
            fileHeader += f' *  <h2><center>&copy; Copyright (c){today.year} CIP United Co.\n'
            fileHeader += """
* All rights reserved.</center></h2>
*
* 
*
******************************************************************************

*/

"""
            fileStr = fileHeader
            fileStr += f'class {modName}_v_reg_test_sequence extends cip_base_sequence;\n\n'
            fileStr += f'{cst_tab_str}`uvm_object_utils({modName}_v_reg_test_sequence)\n\n'
            fileStr += f'{cst_tab_str}function new(string name="{modName}_v_reg_test_sequence");\n'
            fileStr += f'{cst_tab_str}{cst_tab_str}super.new(name);\n{cst_tab_str}endfunction\n\n'
            fileStr += f'{cst_tab_str}virtual task body();\n\n'
            fileStr += f'{cst_tab_str}{cst_tab_str}uvm_reg_hw_reset_seq     reg_rst_seq;\n'
            bAllRegFdHdlPathEmpty = True
            modName_U = modName.upper()
            mod_fd_access_str =''
            modinstCount= len(preip_lst)
            mod_reg_Reset_ignore_str =''
            module_inst = preip_lst[0]
            tab_str = cst_tab_str +cst_tab_str
            if modinstCount > 1 :
                module_index = f'{modName_U}[n]'
                tab_str += cst_tab_str
            else:
                module_index = modName_U
            
            if isinstance(module_inst,St_Peripheral):
                for reg in module_inst.clust_reg_lst:
                    if isinstance(reg,St_Cluster):
                        clu_reset_str,clu_access_str,b =get_sequence_sv_clu(reg,module_index,'',tab_str)
                        mod_reg_Reset_ignore_str += clu_reset_str
                        mod_fd_access_str += clu_access_str
                        if not b:
                            bAllRegFdHdlPathEmpty = False
                        pass
                    elif isinstance(reg,St_Register):
                        # if reg.bVirtual:
                        #     continue
                        if reg.name.upper() == 'RESERVED': #Reserved
                            continue
                        regReset_ignore_str,reg_access_str,bRegFdHdlPathEmpty = get_sequence_sv_reg(reg,module_index,reg.name,tab_str)
                        if regReset_ignore_str:
                            mod_reg_Reset_ignore_str += regReset_ignore_str
                        if reg_access_str:
                            mod_fd_access_str += reg_access_str
                        if not bRegFdHdlPathEmpty:
                            bAllRegFdHdlPathEmpty = False
                        pass
                    pass
                if mod_fd_access_str:
                    mod_fd_access_str += '\n'
                if mod_reg_Reset_ignore_str:
                    mod_reg_Reset_ignore_str += '\n'
                pass

            if not bAllRegFdHdlPathEmpty:
                fileStr += f'{cst_tab_str}{cst_tab_str}uvm_reg_access_seq       reg_access_seq;\n\n'
            fileStr += f'{cst_tab_str}{cst_tab_str}super.body;\n\n'


            fileStr += f'{cst_tab_str}{cst_tab_str}`uvm_info("UVM_SEQ","register reset sequence started",UVM_LOW)\n'
            fileStr += f'{cst_tab_str}{cst_tab_str}reg_rst_seq = new();\n'
            
            if modinstCount > 1 :
                fileStr += f'{cst_tab_str}{cst_tab_str}for (int n=0; n< {modinstCount}; n++) begin\n'
                pass
            fileStr += mod_reg_Reset_ignore_str
            fileStr += f'{tab_str}reg_rst_seq.model = p_sequencer.u_soc_reg_model.{module_index};\n'
            fileStr += f'{tab_str}reg_rst_seq.start(p_sequencer);\n'
            if modinstCount > 1 :
                fileStr += f'{cst_tab_str}{cst_tab_str}end\n\n'
            
            fileStr += f'{cst_tab_str}{cst_tab_str}`uvm_info("UVM_SEQ","register reset sequence finished",UVM_LOW)\n\n'

            if not bAllRegFdHdlPathEmpty:
                fileStr += f'{cst_tab_str}{cst_tab_str}`uvm_info("UVM_SEQ","register access sequence started",UVM_LOW)\n'
                fileStr += f'{cst_tab_str}{cst_tab_str}reg_access_seq = new();\n'
                if modinstCount > 1 :
                    fileStr += f'{cst_tab_str}{cst_tab_str}for (int n=0; n< {modinstCount}; n++) begin\n'
                    pass
                fileStr += mod_fd_access_str      

                fileStr += f'{tab_str}reg_access_seq.model = p_sequencer.u_soc_reg_model.{module_index};\n'
                fileStr += f'{tab_str}reg_access_seq.start(p_sequencer);\n'
  
                if modinstCount > 1 :
                    fileStr += f'{cst_tab_str}{cst_tab_str}end\n\n'

                fileStr += f'{cst_tab_str}{cst_tab_str}`uvm_info("UVM_SEQ","register access sequence finished",UVM_LOW)\n'
            
            fileStr += f'\n{cst_tab_str}endtask: body\n\n'
            fileStr += f'endclass:{modName}_v_reg_test_sequence\n'
            sv_file.write(fileStr)

            return out_file_Pathname
    pass

def output_uvm_sv_moduleFile(preip_lst:list,preip_name:str,version: str):
    preip_inst = None
    if preip_lst:
        preip_inst = preip_lst[0]
        if isinstance(preip_inst,St_Peripheral):
            out_ralf_file_Name = preip_name.lower()
            module_Name = preip_inst.moduleName
            preip_name = preip_name.upper()
            out_file_name = out_ralf_file_Name+'.sv'
            out_file_Pathname = './uvm/'+out_file_name
            with open(out_file_Pathname, 'w+') as out_file:
                fileHeader = f'/**\n * @file    {out_file_name}\n'
                fileHeader += f' * @author  CIP Application Team\n # @brief   {preip_name} Register struct Header File.\n'
                fileHeader += ' *          This file contains:\n #           - Data structures and the address mapping for\n'
                fileHeader += f" *             {preip_name} peripherals\n #           - Including peripheral's registers declarations and bits\n"
                fileHeader += ' *             definition\n'

                # 格式化成2016-03-20 11:45:39形式
                today = date.today()
                fileHeader += f' # @version {version} \n # @date    {today.strftime("%y-%m-%d")}\n'
                fileHeader += """
*
******************************************************************************
* @copyright
*
"""
                fileHeader += f' *  <h2><center>&copy; Copyright (c){today.year} CIP United Co.\n'
                fileHeader += """
* All rights reserved.</center></h2>
*
* 
*
******************************************************************************

*/

"""
                macro_name = f'RAL_MOD_{preip_name}'
                fileHeader += f'`ifndef {macro_name}\n`define {macro_name}\n\n'
                fileHeader += 'import uvm_pkg::*;\n\n'

                block_name, clu_reg_str_info = getCluRegStructInfo_uvm_sv(preip_inst.clust_reg_lst, module_Name,module_Name,0)
                fileHeader += clu_reg_str_info
                # 增加 ral_sys_
                fileHeader += f'\n`endif // {macro_name}' 
                out_file.write(fileHeader)

                return out_file_Pathname
    pass

def output_ralf_moduleFile(preip_lst:list,preip_name:str,version: str):
    preip_inst = None
    if preip_lst:
        preip_inst = preip_lst[0]
        if isinstance(preip_inst,St_Peripheral):
            out_ralf_file_Name = preip_name.lower()
            module_Name = preip_inst.moduleName
            preip_name = preip_name.upper()
            out_file_name = out_ralf_file_Name+'.ralf'
            out_file_Pathname = './uvm/'+out_file_name
            with open(out_file_Pathname, 'w+') as out_file:
                fileHeader = f' # @file    {out_file_name}\n'
                fileHeader += f' # @author  CIP Application Team\n # @brief   {preip_name} Register struct Header File.\n'
                fileHeader += ' #          This file contains:\n #           - Data structures and the address mapping for\n'
                fileHeader += f" #             {preip_name} peripherals\n #           - Including peripheral's registers declarations and bits\n"
                fileHeader += ' #             definition\n'

                # 格式化成2016-03-20 11:45:39形式
                today = date.today()
                fileHeader += f' # @version {version} \n # @date    {today.strftime("%y-%m-%d")}\n'


                fileHeader += f' # <h2><center>&copy; Copyright (c){today.year} CIP United Co.\n'
                fileHeader += """
 # All rights reserved.</center></h2>
 #

"""
                

                clu_reg_str_info = getCluRegStructInfo_Ralf(preip_inst.clust_reg_lst, module_Name,0)
                fileHeader += clu_reg_str_info
          
                
                out_file.write(fileHeader)

                return out_file_Pathname
    pass

def output_C_moduleFile(preip_lst:list, preip_name:str, version:str):
    preip_inst = None
    
    if preip_lst:
        preip_inst = preip_lst[0]
        if isinstance(preip_inst,St_Peripheral):
            out_C_file_Name = preip_name.lower()+'_regs'
            module_Name = preip_inst.headerStructName.upper()
            module_Header= module_Name+'_REGS'
            preip_name = preip_name.upper()
            out_file_name = out_C_file_Name+'.h'
            out_file_Pathname = './reg/'+out_file_name
            with open(out_file_Pathname, 'w+') as out_file:
                fileHeader = f'/**\n * @file    {out_file_name}\n'
                fileHeader += f' * @author  CIP Application Team\n * @brief   {preip_name} Register struct Header File.\n'
                fileHeader += ' *          This file contains:\n *           - Data structures and the address mapping for\n'
                fileHeader += f" *             {preip_name} peripherals\n *           - Including peripheral's registers declarations and bits\n"
                fileHeader += ' *             definition\n'

                # 格式化成2016-03-20 11:45:39形式
                today = date.today()
                fileHeader += f' * @version {version} \n * @date    {today.strftime("%y-%m-%d")}\n'

                fileHeader += """
 *
  ******************************************************************************
 * @copyright
 *
"""
                fileHeader += f' *  <h2><center>&copy; Copyright (c){today.year} CIP United Co.\n'
                fileHeader += """
 * All rights reserved.</center></h2>
 *
 * 
 *
 ******************************************************************************
 */

 /*
 * ****************************************************************************
 * ******** Define to prevent recursive inclusion                  ********
 * ****************************************************************************
 */
#pragma once
"""
                fileHeader += f'#ifndef __{module_Header}_H\n'
                fileHeader += f'#define __{module_Header}_H\n'

                fileHeader += """
#include <stdint.h>

#ifndef   __IO
#define   __IO    volatile
#endif

#ifndef   __I
#define   __I     volatile const
#endif

#ifndef   __O
#define   __O     volatile 
#endif


/*
 * ****************************************************************************
 * ******** Includes                                                   ********
 * ****************************************************************************
 */

/*
 * ****************************************************************************
 * ******** Exported Types                                             ********
 * ****************************************************************************
 */
/*
 * ============================================================================
 * ==            Peripheral registers structures                             ==
 * ============================================================================
 */
/** @addtogroup Peripheral_Registers_Structures
 * @{
 */
#pragma pack(1)

"""
            
                #uint_str = 'uint32_t'

                clu_reg_str_info,fileRegFdOpstr,nLastOffset = getCluRegStructInfo_C(preip_inst.clust_reg_lst, module_Name,0)
                fileHeader += clu_reg_str_info

                fileHeader += '}'+f' {module_Name}_t;\n\n#pragma pack()\n'
                
                fileHeader += """
/**
 * @}
 */

/*
 * ****************************************************************************
 * ******** Exported constants                                         ********
 * ****************************************************************************
 */
/*
 * ============================================================================
 * ==           Device Memory Space Definition                               ==
 * ============================================================================
 */
/** @addtogroup Peripheral_Memory_Map_Definition
 * @{
 */
/* ----------------------------------------------------------------------------
 * -- Memory Space Map
 * ------------------------------------------------------------------------- */

/* ----------------------------------------------------------------------------
 * -- Peripherals Space Map
 * ------------------------------------------------------------------------- */

/* ----------------------------------------------------------------------------
 * -- Special function register map
 * ------------------------------------------------------------------------- */

"""

            
                perip_inst_lst=[]
                for p in  preip_lst:
                    if not p.aliasPeripheral:
                        perip_inst_lst.append(p)
                
                for p in perip_inst_lst:
                    inst_name=p.name.upper()
                    fileHeader+=f'#define  {inst_name}_BASE_ADDR          {inst_name}_BASE\n'

                fileHeader+="""
/**
 * @}
 */

/*
 * ============================================================================
 * ==           Peripheral instance declaration                              ==
 * ============================================================================
 */

/** @addtogroup Peripheral_Instance_Declaration
 * @{
 */

"""
            
                for p in perip_inst_lst:
                    inst_name=p.name.upper()
                    fileHeader+=f'#define  {inst_name}      ( ({module_Name}_t*) {inst_name}_BASE_ADDR )\n'

                fileHeader += """
/**
 * @}
 */


/*
 * ============================================================================
 * ==           Peripheral register bit description                          ==
 * ============================================================================
 */

/** @addtogroup Peripheral_Registers_Bits_Definition
 * @{
 */

/* ----------------------------------------------------------------------------
"""

                mod_reg_str = f' * --     {module_Name}  register bit description\n'
                fileHeader += mod_reg_str
                fileHeader +="""
 * --------------------------------------------------------------------------*/

"""

                fileHeader += fileRegFdOpstr
                fileHeader += f'\n#endif // __{module_Header}_H\n'
                fileHeader += """
/*
 * ****************************************************************************
 * End File
 * ****************************************************************************
 */

"""
                out_file.write(fileHeader)

                return out_file_Pathname

def fieldWriteChk_func(errCount_Write_var:str, str_Tab:str, fd_var:str, module_fd_var:str, strfdMask:str):
    fdWriteCheckstr = ''
    fdWriteCheckstr += f'{str_Tab}{module_fd_var} = {strfdMask};\n'
    fdWriteCheckstr += f'{str_Tab}nRegFdVal = {module_fd_var};\n'
    fdWriteCheckstr += f'{str_Tab}if({module_fd_var} != {strfdMask})\n'
    fdWriteCheckstr += f'{str_Tab}' + '{\n'
    fdWriteCheckstr += f'{str_Tab}{cst_tab_str}Print_time();\n'
    fdWriteCheckstr += f'{str_Tab}{cst_tab_str}Error("Inst_%u # {fd_var}  [0x%X] NOt same as Write [{strfdMask}]! \\n", mi, nRegFdVal);\n'
    fdWriteCheckstr += f'{str_Tab}{cst_tab_str}++{errCount_Write_var};\n'
    fdWriteCheckstr += str_Tab + '}\n'
    return fdWriteCheckstr


def getModule_Clu_FdChkStr(clu:St_Cluster,paretn_clu_name:str,tab_str:str,errCount_var:str, errCount_Write_var:str, modinst_var:str,bModuleLoop: bool = True,nCluevel:int =0):
    filed_resetChk_str = ''
    field_WriteChk_str = ''
    bCluLoop = False
    str_Tab = tab_str
    if bModuleLoop:
        str_Tab += cst_tab_str
    cluName = clu.name
    if clu.dim > 1 and nCluevel == 0 :
        cluName += '[ci]'
        bCluLoop = True
    if paretn_clu_name:
        cluName = paretn_clu_name+'.'+cluName
    for reg in clu.clusters:
        if isinstance(reg,St_Register):
            if bCluLoop:
                str1,str2 = getModule_Reg_FdChkStr(reg,cluName,str_Tab+cst_tab_str,errCount_var,errCount_Write_var,modinst_var,bModuleLoop)
                if str1:
                    filed_resetChk_str += f'{str_Tab}for(int ci=0; ci<{reg.dim}; ++ci)\n'
                    filed_resetChk_str += str_Tab + '{\n'
                    filed_resetChk_str += str1
                    filed_resetChk_str += str_Tab + '}\n'
                if str2:
                    field_WriteChk_str += f'{str_Tab}for(int ci=0; ci<{reg.dim}; ++ci)\n'
                    field_WriteChk_str += str_Tab + '{\n'
                    field_WriteChk_str += str2
                    field_WriteChk_str += str_Tab + '}\n'
                pass
            else:
                if clu.dim >1: 
                    for i in clu.dim:
                        cluNameUse = cluName + f'[{i}]'
                        str1,str2 = getModule_Reg_FdChkStr(reg,cluNameUse,str_Tab,errCount_var,errCount_Write_var,modinst_var,bModuleLoop)
                        filed_resetChk_str += str1
                        field_WriteChk_str += str2
                        pass
                    pass
                else:
                    str1,str2 = getModule_Reg_FdChkStr(reg,cluName,str_Tab,errCount_var,errCount_Write_var,modinst_var,bModuleLoop)
                    filed_resetChk_str += str1
                    field_WriteChk_str += str2
                pass
        elif isinstance(reg,St_Cluster):
            if bCluLoop:
                str1,str2 = getModule_Clu_FdChkStr(reg,cluName,str_Tab+cst_tab_str,errCount_var,errCount_Write_var,modinst_var,bModuleLoop,nCluevel+1)
                if str1:
                    filed_resetChk_str += f'{str_Tab}for(int ci=0; ci<{reg.dim}; ++ci)\n'
                    filed_resetChk_str += str_Tab + '{\n'
                    filed_resetChk_str += str1
                    filed_resetChk_str += str_Tab + '}\n'
                if str2:
                    field_WriteChk_str += f'{str_Tab}for(int ci=0; ci<{reg.dim}; ++ci)\n'
                    field_WriteChk_str += str_Tab + '{\n'
                    field_WriteChk_str += str2
                    field_WriteChk_str += str_Tab + '}\n'
                pass
            else:
                if clu.dim >1: 
                    for i in clu.dim:
                        cluNameUse = cluName + f'[{i}]'
                        str1,str2 = getModule_Clu_FdChkStr(reg,cluNameUse,str_Tab,errCount_var,errCount_Write_var,modinst_var,bModuleLoop,nCluevel+1)
                        filed_resetChk_str += str1
                        field_WriteChk_str += str2
                        pass
                    pass
                else:
                    str1,str2 = getModule_Clu_FdChkStr(reg,cluName,str_Tab,errCount_var,errCount_Write_var,modinst_var,bModuleLoop,nCluevel+1)
                    filed_resetChk_str += str1
                    field_WriteChk_str += str2
                    pass
                pass
            pass
        pass
    return filed_resetChk_str, field_WriteChk_str

def getModule_Reg_FdChkStr_impl(reg:St_Register,reg_Name_use:str,str_Tab:str,modinst_var:str,errCount_var:str, errCount_Write_var:str,bModuleLoop:bool):
    filed_resetChk_str = ''
    field_WriteChk_str = ''
    bRegLoop = reg.dim >1
    for fd in reg.fields:
        fd_name=fd.name.upper()
        if fd_name.startswith('RESERVED'): #reserved
            continue
        if fd.defaultValue == 'X':
            continue
        if reg.nValidFdCount in (0,1):
            reg_fd_var = f'{reg_Name_use}'
            pass
        else:
            reg_fd_var = f'{reg_Name_use}.{fd.name}'
        fd_var = reg_fd_var
        if bRegLoop :
            fd_var = reg_fd_var.replace('[ri]','[%u]')
        module_fd_var = f'{modinst_var}->{reg_fd_var}'
        nBitWid = fd.bitWidth
        fdAccess = fd.access.upper()
        if fdAccess.find('R') != -1:
            filed_resetChk_str += f'{str_Tab}nRegFdVal = {module_fd_var};\n'
            filed_resetChk_str += f'{str_Tab}if(nRegFdVal != {fd.defaultValue})\n'
            filed_resetChk_str += str_Tab + '{\n'
            filed_resetChk_str += f'{str_Tab}{cst_tab_str}Print_time();\n'
            if bRegLoop :
                if bModuleLoop:
                    filed_resetChk_str += f'{str_Tab}{cst_tab_str}Error("Inst_%u # {fd_var}  [0x%X] is NOt same! \\n", mi, ri, nRegFdVal);\n'
                else:
                    filed_resetChk_str += f'{str_Tab}{cst_tab_str}Error("{fd_var}  [0x%X] is NOt same! \\n",ri,nRegFdVal);\n'
                    pass
            else:
                if bModuleLoop :
                    filed_resetChk_str += f'{str_Tab}{cst_tab_str}Error("Inst_%u # {fd_var}  [0x%X] is NOt same! \\n", mi, nRegFdVal);\n'
                else:
                    filed_resetChk_str += f'{str_Tab}{cst_tab_str}Error("{fd_var}  [0x%X] is NOt same! \\n",nRegFdVal);\n'
                    pass
            filed_resetChk_str += f'{str_Tab}{cst_tab_str}++{errCount_var};\n'
            filed_resetChk_str += str_Tab + '}\n'
            if bRegLoop :
                if bModuleLoop:
                    filed_resetChk_str += f'{str_Tab}else\n{str_Tab}{cst_tab_str}Info("Inst_%u # {fd_var} Value is OK. \\n", mi);\n'
                else:
                    filed_resetChk_str += f'{str_Tab}else\n{str_Tab}{cst_tab_str}Info("{fd_var} Value is OK. \\n", mi);\n'
            else:
                if bModuleLoop:
                    filed_resetChk_str += f'{str_Tab}else\n{str_Tab}{cst_tab_str}Info("Inst_%u # {fd_var} Value is OK. \\n");\n'
                else:
                    filed_resetChk_str += f'{str_Tab}else\n{str_Tab}{cst_tab_str}Info("{fd_var} Value is OK. \\n");\n'

        if fdAccess == 'RW':
            if len(fd.enumValues) > 1:
                strfdMask = fd.enumValues[-1].value
                field_WriteChk_str += fieldWriteChk_func(
                    errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

                strfdMask = fd.enumValues[0].value
                field_WriteChk_str+=fieldWriteChk_func(
                    errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)
            else:
                strfdMask = f'{bitWidMask_arr[nBitWid-1]}'
                field_WriteChk_str += fieldWriteChk_func(
                    errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)

                strfdMask = 0
                field_WriteChk_str += fieldWriteChk_func(
                    errCount_Write_var,  str_Tab, fd_var, module_fd_var, strfdMask)
        pass
    return filed_resetChk_str, field_WriteChk_str

def getModule_Reg_FdChkStr(reg:St_Register,paretn_clu_name:str, base_tab_str:str,errCount_var:str, errCount_Write_var:str, modinst_var:str,bModuleLoop: bool = True):
    filed_resetChk_str = ''
    field_WriteChk_str = ''
    str_Tab = base_tab_str
    regName = reg.getRegName().upper()
    if reg.alternateGroupName:
        reg_Name_use = reg.alternateGroupName + '.'+regName
    else:
        reg_Name_use = regName
    if paretn_clu_name:
        reg_Name_use = paretn_clu_name+'.'+reg_Name_use
    if reg.dim > 1 :
        reg_Name_use = reg_Name_use + '[ri]'
        str1,str2 = getModule_Reg_FdChkStr_impl(reg,reg_Name_use,str_Tab + cst_tab_str,modinst_var,errCount_var,errCount_Write_var,bModuleLoop)
        if str1:
            filed_resetChk_str += f'{str_Tab}for(int ri=0; ri<{reg.dim}; ++ri)\n'
            filed_resetChk_str += str_Tab + '{\n'
            filed_resetChk_str += str1
            filed_resetChk_str += str_Tab + '}\n'
        if str2:
            field_WriteChk_str += f'{str_Tab}for(int ri=0; ri<{reg.dim}; ++ri)\n'
            field_WriteChk_str += str_Tab + '{\n'
            field_WriteChk_str += str2
            field_WriteChk_str += str_Tab + '}\n'
        pass
    else:
        str1,str2 = getModule_Reg_FdChkStr_impl(reg,reg_Name_use,str_Tab,modinst_var,errCount_var,errCount_Write_var,bModuleLoop)
        filed_resetChk_str += str1
        field_WriteChk_str += str2
        pass

    return filed_resetChk_str, field_WriteChk_str

def getModule_FdChkStr(mod_inst:St_Peripheral, errCount_var:str, errCount_Write_var:str, modinst_var:str,bModuleLoop: bool = True):
    file_resetChk_str = ''
    field_WriteChk_str = '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
    str_Tab = cst_tab_str
    if bModuleLoop:
        str_Tab += cst_tab_str
    # mod_name = mod_inst.module_name
    for reg in mod_inst.clust_reg_lst:
        if isinstance(reg,St_Cluster):
            if reg.name == 'RESERVED':
                continue
            str1,str2 = getModule_Clu_FdChkStr(reg,'',str_Tab,errCount_var,errCount_Write_var,modinst_var,bModuleLoop)
            file_resetChk_str += str1
            field_WriteChk_str += str2
            pass
        elif isinstance(reg,St_Register):
            if reg.name == 'RESERVED':
                continue
            str1,str2 = getModule_Reg_FdChkStr(reg,'',str_Tab,errCount_var,errCount_Write_var,modinst_var,bModuleLoop)
            file_resetChk_str += str1
            field_WriteChk_str += str2
            pass
        pass

    # fieldWriteCheckstr += '#endif //CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
    return file_resetChk_str, field_WriteChk_str

def output_C_FdValChk_moduleFile(module_inst_list:list, modName:str):
    # print(modName)
    out_C_file_Name=''
    if sys.platform == 'win32':
        dirName = './module_check/'+modName
        if not os.path.exists(dirName):
            os.makedirs(dirName)
        out_C_file_Name = dirName+'/main.c'
    elif sys.platform == 'linux':
        dirName = os.path.join(get_output_c_dir(), modName.lower()+'_reg_c_reg_test')
        out_C_file_Name = dirName+'_main.c'
    with open(out_C_file_Name, 'w+') as out_file:
        fileHeader = """// Autor: Auto generate by python From module excel\n
// Version: 0.0.2 X
// Description : field default value check for module instance \n
// Waring: Do NOT Modify it !
"""

        today = date.today()
        fileHeader += f'// Copyright (C) {today.year} CIP United Co. Ltd.  All Rights Reserved.\n'
        fileHeader += """
#define DEBUG
//#define INFO
#define WARNING
#define NOTICE
#define ERROR
#define PASS
#define FAIL

#define CHECK_MODULE_FIELD_DEFAULT_VALUE

#define CHECK_MOUDLE_FIELD_WRITE_VALUE

#include <stdio.h>
#include <time.h>

// void getCurrentTimeStr(char* const szTimeBuf,int nBufSize ){
//     struct timespec ts;
//     timespec_get(&ts, TIME_UTC);
//     struct tm * lct = localtime(&ts.tv_sec);
//     sprintf_s(szTimeBuf,nBufSize,"time: %02d %02d:%02d:%02d [%09ld]", lct->tm_mday,lct->tm_hour,lct->tm_min,lct->tm_sec,ts.tv_nsec);
// }

// include "log.h"
// include "pll.h"

"""

        for perip in module_inst_list:
            perip_str = f'#define {perip.name}_BASE'.ljust(cst_Perip_SpaceSize)+f'({perip.getDirectAddrStr()})\n'
            fileHeader += perip_str

        filebodystr = f'\n#include "../../reg/{modName.lower()}_regs.h"\n'
        filebodystr += """
int main()
{
    //printf("enter main.\\n");
    uAptiv_clk_init();
"""
        
        mod_inst = module_inst_list[0]
        module_Name = mod_inst.headerStructName.upper()
        module_st_name = f'{module_Name}_t'
        filebodystr += f'{cst_tab_str}printf("After clock switch, Now Check Module: {modName}.\\n");\n'
        filebodystr += f'{cst_tab_str}unsigned int nRegFdVal = 0;\n'
        filebodystr += f'{cst_tab_str}unsigned int nTotalErr = 0;\n'

        mod_count = len(module_inst_list)
        if mod_count > 1:
            filebodystr += f'{cst_tab_str}{module_st_name} * module_inst[{mod_count}] = ' + \
                '{'+f'{mod_inst.name}'
            for i in range(1, mod_count):
                filebodystr += f'\n{cst_tab_str}{cst_tab_str},{module_inst_list[i].name}'
            filebodystr += '};\n\n'
            filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
            filebodystr += f'{cst_tab_str}unsigned int nErrCount_default[{mod_count}] = '+'{0};\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
            filebodystr += f'{cst_tab_str}unsigned int nErrCount_wirte[{mod_count}] = '+'{0};\n'
            filebodystr += '#endif // CHECK_MOUDLE_FIELD_WRITE_VALUE\n\n'

            filebodystr += f'{cst_tab_str}for(int mi = 0; mi < {mod_count}; ++mi)\n'
            filebodystr += cst_tab_str + '{\n'
            filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
            modinst_var = 'module_inst[mi]'
            errCount_var = 'nErrCount_default[mi]'
            errCount_write_var = 'nErrCount_wirte[mi]'
            str1, str2 = getModule_FdChkStr(mod_inst, errCount_var, errCount_write_var, modinst_var)
            filebodystr += str1
            filebodystr += f'{cst_tab_str}{cst_tab_str}if(nErrCount_default[mi])\n'
            filebodystr += cst_tab_str+cst_tab_str+'{\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}Error("Inst_%u def-Vals have [%u] fds NOT Same!\\n", mi, nErrCount_default[mi]);\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}nTotalErr += nErrCount_default[mi];\n'
            filebodystr += cst_tab_str+cst_tab_str+'}\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}else\n{cst_tab_str}{cst_tab_str}{cst_tab_str}Notice("Inst_%u def-Vals are OK!\\n", mi);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += str2
            filebodystr += f'{cst_tab_str}{cst_tab_str}if(nErrCount_wirte[mi])\n'
            filebodystr += cst_tab_str+cst_tab_str+'{\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}Error("Inst_%u write-Vals have [%u] fds NOT Same!\\n", mi, nErrCount_wirte[mi]);\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}nTotalErr += nErrCount_wirte[mi];\n'
            filebodystr += cst_tab_str+cst_tab_str+'}\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}else\n{cst_tab_str}{cst_tab_str}{cst_tab_str}Notice("Inst_%u write-Vals are OK!\\n", mi);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'
            filebodystr += cst_tab_str+'}\n'
        elif mod_count == 1:
            filebodystr += f'{cst_tab_str}{module_st_name} * module_inst = {mod_inst.name} ;\n'
            filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
            filebodystr += f'{cst_tab_str}unsigned int nErrCount_default = 0;\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
            filebodystr += f'{cst_tab_str}unsigned int nErrCount_wirte = 0;\n'
            filebodystr += '#endif // CHECK_MOUDLE_FIELD_WRITE_VALUE\n\n'

            modinst_var = 'module_inst'
            errCount_var = 'nErrCount_default'
            errCount_write_var = 'nErrCount_wirte'
            # filebodystr += getModuleFdStr(mod_inst,
            #                               errCount_var, modinst_var, False)
            str1, str2 = getModule_FdChkStr(mod_inst, errCount_var, errCount_write_var, modinst_var, False)

            filebodystr += str1

            filebodystr += f'{cst_tab_str}{cst_tab_str}if(nErrCount_default)\n'
            filebodystr += cst_tab_str+cst_tab_str+'{\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}Error("Inst_%u def-Vals have [%u] fds NOT Same!\\n", mi, nErrCount_default);\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}nTotalErr += nErrCount_default;\n'
            filebodystr += cst_tab_str+cst_tab_str+'}\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}else\n{cst_tab_str}{cst_tab_str}{cst_tab_str}Notice("Inst_%u def-Vals are OK!\\n", mi);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

            filebodystr += str2
            filebodystr += f'{cst_tab_str}{cst_tab_str}if(nErrCount_wirte)\n'
            filebodystr += cst_tab_str+cst_tab_str+'{\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}Error("Inst_%u write-Vals have [%u] fds NOT Same!\\n", mi, nErrCount_wirte);\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}{cst_tab_str}nTotalErr += nErrCount_wirte;\n'
            filebodystr += cst_tab_str+cst_tab_str+'}\n'
            filebodystr += f'{cst_tab_str}{cst_tab_str}else\n{cst_tab_str}{cst_tab_str}{cst_tab_str}Notice("Inst_%u write-Vals are OK!\\n", mi);\n'
            filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

       
        filebodystr += f'{cst_tab_str}if(nTotalErr == 0)\n'
        filebodystr += f'{cst_tab_str}{cst_tab_str}Pass("{modName} Vals OK!\\n");\n'
        filebodystr += f'{cst_tab_str}else\n{cst_tab_str}{cst_tab_str}Fail("{modName} Vals Not OK!\\n");\n'
        filebodystr += f'\n{cst_tab_str}return 0;\n'+'}\n'
        out_file.write(fileHeader)
        out_file.write(filebodystr)
        out_file.close()

        return out_C_file_Name


def getHexStr(var: int|str):
    if isinstance(var,int):
        e_val = hex(var)
        hxstr=hex(var)
        return hxstr[2:].upper()
    elif isinstance(var,str):
        var_up = var.upper()
        if var_up.startswith('0X'):
            return var_up[2:]


def getRegFieldInfo_uvm_sv(clu_reg: St_Register,moduleName:str,cluster_level: int = 0):
    regName = clu_reg.name.upper()
    cls_reg_str = ''
    cls_constraint_str =''
    cls_build_str = ''
    cls_reg_name = ''
    block_build_str = ''
    if regName != 'RESERVED': #Reserved
        reg_name = f'{moduleName}_{clu_reg.name}'
        cls_reg_name = f'ral_reg_{reg_name}'
        cls_reg_str = f'\nclass {cls_reg_name} extends uvm_reg;\n'
        reg_tab_str =   cst_tab_str
        field_tab_str =  cst_tab_str + cst_tab_str
        hdl_path = ''
        bfd_individually_accessible = 0
        nUsefulFd = 0
        for fd in clu_reg.fields:
            if isinstance(fd,St_Field):
                fdName = fd.name.upper()
                if fdName != 'RESERVED':
                    nUsefulFd += 1

        if nUsefulFd == 1:
            bfd_individually_accessible = 1
        for fd in clu_reg.fields:
            if isinstance(fd,St_Field):
                fdName = fd.name.upper()
                if fdName != 'RESERVED':
                    fdAccess = fd.access.lower()
                    if fdAccess == 'r' or fdAccess == 'w':
                        fdAccess += 'o'
                    if fdAccess.find('w') != -1:
                        cls_reg_str += reg_tab_str + f'rand uvm_reg_field {fd.name};\n'
                    else:
                        cls_reg_str += reg_tab_str + f'uvm_reg_field {fd.name};\n'
                    cls_build_str += field_tab_str + f'this.{fd.name} = uvm_reg_field::type_id::create("{fd.name}",,get_full_name());\n'
                    cls_build_str += field_tab_str + f'this.{fd.name}.configure(this, {fd.bitWidth}, {fd.bitOffset}, "{fdAccess.upper()}", 0, {fd.bitWidth}\'h{getHexStr(fd.defaultValue)}, 1, 1, {bfd_individually_accessible});\n'

                    cls_constraint_str += reg_tab_str + f'constraint {clu_reg.name}_cst_{fd.name}'+' {\n'
                    constraint_str = ''
                    enum_str =''
                    enum_val_str =''
                    if fd.writeConstraint == 'enumerated':
                        if fd.enumValues:
                            enum_str += field_tab_str  + 'enum {\n'
                            bFirstEnum = True
                            for e in fd.enumValues:
                                enum_str += field_tab_str + cst_tab_str
                                if not bFirstEnum:
                                    enum_str +=','
                                    enum_val_str += ','
                                    pass
                                if isinstance(e,St_Enum_Val):
                                    enum_str += f'{e.name} = {e.value}\n'
                                    # e_val = e.value
                                    e_val = getHexStr(e.value)
                                    enum_val_str += f'\'h{e_val}'
                                    pass 
                                bFirstEnum = False
                                pass
                            enum_str += field_tab_str + cst_tab_str + '}\n'
                            if len(fd.enumValues) == 1:
                                constraint_str = field_tab_str  + f'{fd.name}.value == {enum_val_str}'+';\n'
                                pass
                            else:
                                constraint_str = field_tab_str  + f'{fd.name}'+ '.value inside {'+f'{enum_val_str}'+'};\n'
                            pass
                        pass
                    elif fd.writeConstraint == 'range':
                        if fd.range_max and fd.range_min:
                            constraint_str = field_tab_str  + f'{fd.name}' + '.value inside { ['+f'\'h{getHexStr(fd.range_min)}:\'h{getHexStr(fd.range_max)}'+'] };\n'
                            pass
                        pass
                    cls_constraint_str += constraint_str
                    cls_constraint_str += reg_tab_str + '}\n'
                    if fd.hdl_path:
                        if hdl_path:
                            hdl_path += ',\n'
                        hdl_path += field_tab_str+'\'{"'+f'{fd.hdl_path}", {fd.bitOffset}, {fd.bitWidth}' + '}'
                    # '{"U_DW_apb_uart_regfile.rbr[7:0]", 0, 8}
                    pass
                pass
        cls_reg_str += cls_constraint_str
        cls_reg_str += reg_tab_str + f'function new(string name = "{reg_name}");\n'    
        cls_reg_str += field_tab_str + f'super.new(name, {clu_reg.size},build_coverage(UVM_NO_COVERAGE));\n'
        cls_reg_str += reg_tab_str + 'endfunction: new\n'
        cls_reg_str += reg_tab_str+ 'virtual function void build();\n'
        cls_reg_str += cls_build_str
        cls_reg_str += reg_tab_str+ 'endfunction: build\n\n'
        cls_reg_str += reg_tab_str+ f'`uvm_object_utils({cls_reg_name})\n\n'
        cls_reg_str +=  f'endclass : {cls_reg_name}\n'

        reg_Access = clu_reg.access.upper()
        if reg_Access == 'R' or reg_Access == 'W':
            reg_Access += 'O'
        if clu_reg.dim > 1:
            block_build_str += field_tab_str + f'foreach (this.{regName}[i]) begin\n'
            block_build_str += field_tab_str + cst_tab_str + f'this.{regName}[i] = {cls_reg_name}::type_id::create("{regName}",,get_full_name());\n'
            if cluster_level:
                block_build_str += field_tab_str + cst_tab_str + f'this.{regName}[i].configure(get_block(), this, "{clu_reg.hdl_path}");\n' 
            else:
                block_build_str += field_tab_str + cst_tab_str + f'this.{regName}[i].configure(this, null, "{clu_reg.hdl_path}");\n'
            block_build_str += field_tab_str + cst_tab_str + f'this.{regName}[i].build();\n'
            if hdl_path:
                block_build_str += field_tab_str + cst_tab_str + f'this.{regName}[i].add_hdl_path('+'\'{\n'
                block_build_str += hdl_path + '\n'
                block_build_str += field_tab_str + cst_tab_str + '});\n' 
            if cluster_level == 0 :
                block_build_str += field_tab_str + cst_tab_str + f'this.default_map.add_reg(this.{regName}[i], `UVM_REG_ADDR_WIDTH\'h{getHexStr(clu_reg.addressOffset)}, "{reg_Access}", 0);\n'
            else:
                block_build_str += field_tab_str + cst_tab_str + f'this.get_topblock_map().add_reg(this.{regName}[i], `UVM_REG_ADDR_WIDTH\'h{getHexStr(clu_reg.addressOffset)}, "{reg_Access}", 0);\n'
                pass
            block_build_str += field_tab_str + 'end\n\n'               
        else:
            block_build_str += field_tab_str + f'this.{regName} = {cls_reg_name}::type_id::create("{regName}",,get_full_name());\n'
            if cluster_level:
                block_build_str += field_tab_str +  f'this.{regName}.configure(get_block(), this, "{clu_reg.hdl_path}");\n' 
            else:
                block_build_str += field_tab_str + f'this.{regName}.configure(this, null, "{clu_reg.hdl_path}");\n'
            block_build_str += field_tab_str + f'this.{regName}.build();\n'
            if hdl_path:
                block_build_str += field_tab_str + f'this.{regName}.add_hdl_path('+'\'{\n'
                block_build_str += hdl_path  + '\n'
                block_build_str += field_tab_str + '});\n' 
            if cluster_level ==0 :
                block_build_str += field_tab_str + f'this.default_map.add_reg(this.{regName}, `UVM_REG_ADDR_WIDTH\'h{getHexStr(clu_reg.addressOffset)}, "{reg_Access}", 0);\n\n'
            else:
                block_build_str += field_tab_str +  f'this.get_topblock_map().add_reg(this.{regName}, `UVM_REG_ADDR_WIDTH\'h{getHexStr(clu_reg.addressOffset)}, "{reg_Access}", 0);\n'
    
    return cls_reg_name,block_build_str,cls_reg_str

def getRegFieldInfo_Ralf(tab_str:str, clu_reg: St_Register,baseOffset:int):
    regName = clu_reg.name.upper()
    fileHeader = ''
    if regName != 'RESERVED': #Reserved
        fileHeader = tab_str+'register ' + clu_reg.name
        if clu_reg.dim >1 :
            fileHeader += f'[{clu_reg.dim}]'
        offset_val = getIntValFromHexString(clu_reg.addressOffset) - baseOffset
        fileHeader += f' @{offset_val} ' +'{\n'
        nbytes = ubitSize_bytes_dict[clu_reg.size]
        reg_tab_str =  tab_str + cst_tab_str
        fileHeader += reg_tab_str + f'bytes  {nbytes};\n'
        field_tab_str = tab_str + cst_tab_str + cst_tab_str
        for fd in clu_reg.fields:
            if isinstance(fd,St_Field):
                fdName = fd.name.upper()
                if fdName != 'RESERVED':
                    fileHeader += reg_tab_str + f'field {fd.name} '
                    if fd.hdl_path:
                        fileHeader +=f' ({fd.hdl_path}) '
                    fileHeader +=f' @{fd.bitOffset} ' + '{\n'
                    fileHeader += field_tab_str + f'bits {fd.bitWidth};\n'
                    # if fd.access == 'R':
                    #     pass
                    fileHeader += field_tab_str + f'reset {fd.defaultValue};\n'
                    fdAccess = fd.access.lower()
                    if fdAccess == 'r' or fdAccess == 'w':
                        fdAccess += 'o'
                    fileHeader += field_tab_str + f'access {fdAccess};\n'
                    fileHeader += field_tab_str + f'constraint c_st_{regName}_{fd.name}'+ ' {\n'
                    constraint_str = ''
                    enum_str =''
                    enum_val_str =''
                    if fd.writeConstraint == 'enumerated':
                        if fd.enumValues:
                            enum_str += field_tab_str  + 'enum {\n'
                            bFirstEnum = True
                            for e in fd.enumValues:
                                enum_str += field_tab_str + cst_tab_str
                                if not bFirstEnum:
                                    enum_str +=','
                                    enum_val_str += ','
                                    pass
                                if isinstance(e,St_Enum_Val):
                                    enum_str += f'{e.name} = {e.value}\n'
                                    # e_val = e.value
                                    e_val = getHexStr(e.value)
                                    enum_val_str += f'\'h{e_val}'
                                    pass 
                                bFirstEnum = False
                                pass
                            enum_str += field_tab_str + cst_tab_str + '}\n'
                            if len(fd.enumValues) == 1:
                                constraint_str = field_tab_str + cst_tab_str + 'value == '+f'{enum_val_str}'+';\n'
                                pass
                            else:
                                constraint_str = field_tab_str + cst_tab_str + 'value inside {'+f'{enum_val_str}'+'};\n'
                            pass
                        pass
                    elif fd.writeConstraint == 'range':
                        if fd.range_max and fd.range_min:
                            constraint_str = field_tab_str + cst_tab_str + 'value inside { ['+f'\'h{getHexStr(fd.range_min)}:\'h{getHexStr(fd.range_max)}'+'] };\n'
                            pass
                        pass
                    fileHeader += constraint_str
                    fileHeader += field_tab_str + '};\n'
                    fileHeader += reg_tab_str + '}; '
                    if fd.description:
                        fileHeader += f' #{fd.description}'
                    fileHeader += '\n'
                    pass
                pass
        fileHeader += tab_str +'}; '
        if clu_reg.description:
            fileHeader += f'# {clu_reg.description}'
        fileHeader += '\n'
    return fileHeader


def out_sys_uvm_sv(st_dev: St_Device):
    #基于domain
    pass    

def getCluRegStructInfo_uvm_sv(clust_reg_lst:list, module_name:str,clu_name:str, nChild_level:int):
    clu_reg_str = ''
    clu_blcok_str = ''
    clu_build_str = ''
    fileHeader = ''
    cls_reg_name = ''
    cluster_fun_str = ''
    for clu_reg in clust_reg_lst:
        if isinstance(clu_reg, St_Cluster):
            regfile_Name,clu_reg_child_str_info = getCluRegStructInfo_uvm_sv(clu_reg.clusters,module_name,clu_reg.name,nChild_level+1)
            fileHeader += clu_reg_child_str_info
            clu_build_str += cst_tab_str + cst_tab_str + f'this.{clu_reg.name} = ral_regfile_{module_name}_{clu_reg.name}::type_id::create("{clu_reg.name}",,get_full_name());\n'
            if nChild_level == 0:
                clu_build_str += cst_tab_str + cst_tab_str + f'this.{clu_reg.name}.configure(this,null, "");\n'
            else:
                clu_build_str += cst_tab_str + cst_tab_str + f'this.{clu_reg.name}.configure(get_block(), this, "");\n'

            clu_build_str += cst_tab_str + cst_tab_str + f'this.{clu_reg.name}.build();\n\n'

            if clu_reg.dim > 1:
                clu_blcok_str += cst_tab_str+ f'rand {regfile_Name} {clu_reg.name}[{clu_reg.dim}];\n'
            else:
                clu_blcok_str += cst_tab_str+ f'rand {regfile_Name} {clu_reg.name};\n'


            pass
        elif isinstance(clu_reg, St_Register):
            reg_name,block_build_str,regFdInfo = getRegFieldInfo_uvm_sv(clu_reg,module_name,nChild_level)
            if regFdInfo:
                fileHeader += regFdInfo 
                clu_build_str += block_build_str 
                if clu_reg.dim > 1:
                    clu_blcok_str += cst_tab_str+ f'rand {reg_name} {clu_reg.name}[{clu_reg.dim}];\n'
                else:
                    clu_blcok_str += cst_tab_str+ f'rand {reg_name} {clu_reg.name};\n'

    if nChild_level == 0:
        cluster_fun_str += '\n' + cst_tab_str + f'function new(string name = "{clu_name}");\n' 
        cluster_fun_str += cst_tab_str + cst_tab_str + 'super.new(name, build_coverage(UVM_NO_COVERAGE));\n'   
        cluster_fun_str += cst_tab_str + 'endfunction: new\n'
        cls_reg_name = f'ral_block_{module_name}'
        clu_reg_str = f'\nclass {cls_reg_name} extends uvm_reg_block;\n'
        clu_build_str = cst_tab_str + cst_tab_str + 'this.default_map = create_map("", 0, 4, UVM_LITTLE_ENDIAN, 0);\n' + clu_build_str
    else:
        cluster_fun_str += '\n' + cst_tab_str + f'function new(string name = "{module_name}_{clu_name}");\n' 
        cluster_fun_str += cst_tab_str + cst_tab_str + 'super.new(name);\n'   
        cluster_fun_str += cst_tab_str + 'endfunction: new\n'
        cluster_fun_str += '\n' + cst_tab_str + 'function uvm_reg_map get_topblock_map();\n'
        cluster_fun_str += cst_tab_str + cst_tab_str + f'return this.get_parent().default_map;\n'
        cluster_fun_str += cst_tab_str + 'endfunction : get_topblock_map\n\n'
        cls_reg_name = f'ral_regfile_{module_name}_{clu_name}'
        clu_reg_str = f'\nclass {cls_reg_name} extends uvm_reg_file;\n'
        pass
    fileHeader += clu_reg_str 
    fileHeader += clu_blcok_str
    
    fileHeader += cluster_fun_str

    fileHeader += '\n'+ cst_tab_str + 'virtual function void build();\n' 

    fileHeader += clu_build_str  
    fileHeader += cst_tab_str + 'endfunction : build\n\n'

    fileHeader += cst_tab_str + f'`uvm_object_utils({cls_reg_name})\n'
    fileHeader +=  f'endclass : {cls_reg_name}\n'

    return cls_reg_name, fileHeader



def getCluRegStructInfo_Ralf(clust_reg_lst:list, struct_Name:str, nChild_level:int = 0, baseOffset:int = 0):
    cst_newLine_tab_str= ''
    for l in range(nChild_level):
        cst_newLine_tab_str += cst_tab_str
    if nChild_level == 0:
        fileHeader = cst_newLine_tab_str+f'block {struct_Name}'+ ' {\n' + cst_newLine_tab_str + cst_tab_str +'bytes 4;\n'
    else:
        fileHeader = cst_newLine_tab_str+f'regfile {struct_Name} @{baseOffset}'+ ' {\n'
    newLine_tab_str = cst_newLine_tab_str + cst_tab_str
    
    for clu_reg in clust_reg_lst:
        if isinstance(clu_reg, St_Cluster):
            regFileName = clu_reg.name
            if clu_reg.dim > 1:
                regFileName += f'[clu_reg.dim]'
            clu_reg_str_info = getCluRegStructInfo_Ralf(clu_reg.clusters,regFileName,nChild_level+1,getIntValFromHexString(clu_reg.addressOffset))
            fileHeader += clu_reg_str_info
            fileHeader += newLine_tab_str+'} ;\n'
            pass
        elif isinstance(clu_reg, St_Register):
            regFdInfo = getRegFieldInfo_Ralf(newLine_tab_str,clu_reg,baseOffset)
            fileHeader += regFdInfo        
    if nChild_level == 0:
        fileHeader += cst_newLine_tab_str + '} ; \n'
    return fileHeader


def getCluRegStructInfo_C(clust_reg_lst:list, struct_name:str, nPLastOffset:int,nChild_level:int = 0):
    cst_newLine_tab_str= ''
    for l in range(nChild_level):
        cst_newLine_tab_str += cst_tab_str
    if nChild_level == 0:
        fileHeader = cst_newLine_tab_str+'typedef struct {\n'
    else:
        fileHeader = cst_newLine_tab_str+'struct {\n'
    fileRegFdOpstr=''
    curGroupName = ''
    newLine_tab_str = cst_newLine_tab_str + cst_tab_str
    nRegReservedIndex = 0
    nLastOffset = nPLastOffset
    for clu_reg in clust_reg_lst:
        if isinstance(clu_reg, St_Cluster):
            if curGroupName:
                fileHeader += newLine_tab_str+'} '+f'{curGroupName};\n'
                curGroupName = ''
            clu_reg_str_info, clu_reg_Op_str,nLastOffset = getCluRegStructInfo_C(clu_reg.clusters,struct_name,nLastOffset,nChild_level+1)
            fileHeader += clu_reg_str_info
            fileRegFdOpstr += clu_reg_Op_str
            fileHeader += newLine_tab_str+'} '+f'{clu_reg.name}'
            if clu_reg.dim >1:
                fileHeader += f'[{clu_reg.dim}]'
            fileHeader += ';\n'
            pass
        elif isinstance(clu_reg, St_Register):
            if clu_reg.alternateGroupName:
                union_str = ''
                if curGroupName != clu_reg.alternateGroupName:
                    if curGroupName:
                        fileHeader += newLine_tab_str+'} '+f'{curGroupName};\n'
                    union_str = newLine_tab_str+'union {\n'
                    curGroupName = clu_reg.alternateGroupName
                offsetInfo,regFdInfo,regFdOpstr,nRegReservedIndex,nLastOffset = getRegFieldInfo_C(newLine_tab_str, clu_reg,struct_name,nRegReservedIndex,nLastOffset)
                fileHeader += offsetInfo
                fileHeader += union_str
                fileHeader += regFdInfo
                fileRegFdOpstr += regFdOpstr
            else:
                if curGroupName:
                    fileHeader +=  newLine_tab_str+'} '+f'{curGroupName};\n'
                    curGroupName = ''
                offsetInfo,regFdInfo,regFdOpstr,nRegReservedIndex,nLastOffset = getRegFieldInfo_C(cst_newLine_tab_str, clu_reg,struct_name,nRegReservedIndex,nLastOffset)
                fileHeader += offsetInfo
                fileHeader += regFdInfo
                fileRegFdOpstr += regFdOpstr
    return fileHeader, fileRegFdOpstr,nLastOffset



def getRegFieldInfo_C(tab_str:str, clu_reg: St_Register ,moduleName:str,nRegReservedIndex:int, nLastOffset:int):
    # 需要增加 dim 部分的处理逻辑
    fileHeader = ''
    retOpstr = f'/**\n * @name {clu_reg.name} - {clu_reg.description}, Offset: {clu_reg.addressOffset}\n * @'+'{\n */\n'

    if clu_reg.size in uint_dict:
        uint_str = uint_dict[clu_reg.size]

    row_tab_str = tab_str+ cst_tab_str
    row_fd_tab_str = row_tab_str + cst_tab_str

    offsetfileHeader = ''
    reg_offset = clu_reg.addressOffset
    if isHexString(reg_offset):
        reg_offset = getIntValFromHexString(reg_offset)
    sizeinfo = f'size:  {math.floor(clu_reg.size/8)}'
    if clu_reg.dim:
        sizeinfo += f', dim: {clu_reg.dim} * {clu_reg.dimIncrement}'
    print(f'reg: {clu_reg.name}, reg_offset: {reg_offset}, {sizeinfo} , LastOffset: {nLastOffset}')
    if reg_offset > nLastOffset:
        nNeedReserved =reg_offset - nLastOffset
        if nNeedReserved >1:
            offsetfileHeader = cst_tab_str +  f'uint8_t nReg_Reserved{nRegReservedIndex}[{nNeedReserved}];'
        else:
            offsetfileHeader = cst_tab_str +  f'uint8_t nReg_Reserved{nRegReservedIndex};'
        offsetfileHeader = offsetfileHeader.ljust(cst_Reg_SpaceSize) + '//Add Reserved for offset  \n'
        nRegReservedIndex += 1
        pass

    if clu_reg.access in accessDict:
        accChar = accessDict[clu_reg.access]
        fileHeader += tab_str+cst_tab_str+ f'{accChar} '
    nAddReservedRegBytes = 0
    if clu_reg.dim:
        if isinstance(clu_reg.dimIncrement,int):
            nIncrement =  clu_reg.dimIncrement*8
            if clu_reg.size < nIncrement:
                nAddReservedRegBytes = math.floor((clu_reg.dimIncrement-clu_reg.size/8))
        nLastOffset = reg_offset + clu_reg.dimIncrement * clu_reg.dim
    else:
        nLastOffset = reg_offset + math.floor(clu_reg.size/8)

    regName = clu_reg.getRegName().upper()
    space_str = cst_tab_str.ljust(cst_Reg_SpaceSize)
    lastRow_str = ''
    if regName == 'RESERVED': #Reserved
        lastRow_str = fileHeader + f'{uint_str} {regName}_{nRegReservedIndex}'
        fileHeader = ''
        if nAddReservedRegBytes > 0:
            struct_str =  row_tab_str + 'struct {\n'
            struct_str += cst_tab_str + lastRow_str +';\n'
            row_str =  row_fd_tab_str +  f'uint8_t nReg_Reserved{nRegReservedIndex};'
            if nAddReservedRegBytes > 1:
                row_str = row_fd_tab_str +  f'uint8_t nReg_Reserved{nRegReservedIndex}[{nAddReservedRegBytes}];'
            row_str = row_str.ljust(cst_Reg_SpaceSize)
            row_str += '//Reserved for dim increment \n'
            struct_str += row_str 
            fileHeader = struct_str
            lastRow_str = row_tab_str+'} Dim_'+f'{regName}_{nRegReservedIndex}'      
            pass 
        if clu_reg.dim:
            lastRow_str += f'[{clu_reg.dim}];'
        else:
            lastRow_str += ';'
        lastRow_str = lastRow_str.ljust(cst_Reg_SpaceSize)
        fileHeader += lastRow_str + f'/*!< Offset: {clu_reg.addressOffset} ({clu_reg.access}),  {clu_reg.description} */\n'
        lastRow_str = ''
        nRegReservedIndex += 1
        pass
    else:
        regFdStr=''
        nValidFdCount = clu_reg.nValidFdCount

        if nValidFdCount in (0,1) :
            lastRow_str = fileHeader + uint_str
            fileHeader = ''
        else:
            if clu_reg.fields:
                regFdStr = 'struct {\n'
                curBitPos=0
                reservedindex =0
                for fd in clu_reg.fields:
                    if isinstance(fd,St_Field):
                        row_str=''
                        fd_Op_str = ''
                        fd_head_str = fd_body_str = ''

                        if fd.bitOffset > curBitPos:
                            #添加reserved
                            row_str+=f'{row_fd_tab_str}{uint_str} reserved{reservedindex}: {curBitPos - fd.bitOffset};'
                            reservedindex += 1
                            nValidFdCount += 1

                        curBitPos = fd.bitOffset + fd.bitWidth
                        
                        if fd.name == 'RESERVED':
                            if curBitPos != clu_reg.size:
                                #最上面的保留字段，不生成
                                row_str+=f'{row_fd_tab_str}{uint_str} reserved{reservedindex}: {fd.bitWidth};'
                                reservedindex+=1
                            else:
                                continue
                        else:
                            nValidFdCount += 1
                            row_str+=f'{row_fd_tab_str}{uint_str} {fd.name}: {fd.bitWidth};'

                            endBit = fd.bitWidth + fd.bitOffset - 1
                            fd_head_str = f'/** Bit[{endBit}:{fd.bitOffset}] {fd.name}  - {fd.access}, {fd.description}\n'
                            if fd.writeConstraint == 'range':
                                fd_head_str += f' * Range: ( {fd.range_min} --{fd.range_max})\n'
                                pass

                            regFdName = f'{moduleName}_{clu_reg.name}_{fd.name}'
                            regFdPos = f'{regFdName}_POS'
                            regFdOpStr = f'#define {regFdPos}'
                            regFdOpStr = regFdOpStr.ljust(cst_RegField_SpaceSize)
                            regFdOpStr += f'{fd.bitOffset}U\n'
                            fd_body_str += regFdOpStr
                            regFdMsk = f'{regFdName}_MSK'
                            regFdOpStr = f'#define {regFdMsk}'
                            regFdOpStr = regFdOpStr.ljust(cst_RegField_SpaceSize)
                            markVal = bitWidMask_arr[fd.bitWidth-1]
                            regFdOpStr += f'(({uint_str}) {markVal} << {regFdPos})\n'
                            fd_body_str += regFdOpStr
                            if fd.enumValues:
                                #仅定义 enum 
                                for e in fd.enumValues:
                                    enumName=e.name.upper()
                                    regFdOpStr = f'#define {regFdName}_{enumName}'
                                    regFdOpStr = regFdOpStr.ljust(cst_RegField_SpaceSize)
                                    regFdOpStr += f'({e.value}U << {regFdPos})\n'
                                    fd_body_str += regFdOpStr

                                    fd_head_str += f' * - {e.value} : {e.desc}\n'
                                pass
                            else:
                                regFdOpStr = f'#define {regFdName}_GET(val)'
                                regFdOpStr = regFdOpStr.ljust(cst_RegField_SpaceSize)
                                regFdOpStr += f'(({uint_str}) ((val) & {regFdMsk}) >> {regFdPos})\n'
                                fd_body_str += regFdOpStr
                                if fd.access.find('W') != -1:
                                    regFdOpStr = f'#define {regFdName}_SET(val)'
                                    regFdOpStr = regFdOpStr.ljust(cst_RegField_SpaceSize)
                                    regFdOpStr += f'(({uint_str}) ((val) & {markVal}) << {regFdPos})\n'
                                    fd_body_str += regFdOpStr
                            fd_head_str += ' */\n'

                        fd_Op_str = fd_head_str + fd_body_str +'\n'
                        retOpstr += fd_Op_str
                        row_str = row_str.ljust(cst_RegField_SpaceSize)
                        row_str += f'/*!< bitOffset: {fd.bitOffset} ({fd.access}), {fd.description} */\n'
                        regFdStr += row_str
                        pass
                    pass
                pass
            # regFdStr += row_tab_str+'}'
            fileHeader += regFdStr
            lastRow_str = row_tab_str+'}'
        lastRow_str += ' ' + clu_reg.getRegName()
        if nAddReservedRegBytes > 0:
            struct_str =  row_tab_str + 'struct {\n'
            struct_str += cst_tab_str + fileHeader +';\n'
            if nAddReservedRegBytes > 1:
                struct_str += row_fd_tab_str +  f'uint8_t nReg_Reserved[{nAddReservedRegBytes}];  //Reserved for dim increment \n'
            else:
                struct_str += row_fd_tab_str +  f'uint8_t nReg_Reserved;  //Reserved for dim increment \n'
            # struct_str += row_tab_str+'} Dim_'+f'{regName}_{nRegReservedIndex}' 
            fileHeader = struct_str
            lastRow_str = row_tab_str+'} Dim_'+f'{regName}_{nRegReservedIndex}' 
            pass
        if clu_reg.dim:
            lastRow_str += f'[{clu_reg.dim}];'
        else:
            lastRow_str += ';'
        lastRow_str = lastRow_str.ljust(cst_Reg_SpaceSize)
        fileHeader += lastRow_str + f'/*!< Offset: {clu_reg.addressOffset} ({clu_reg.access}),  {clu_reg.description} */\n'
        lastRow_str = ''
        pass

    retOpstr += '/**\n * @}\n*/\n\n'
    
    return  offsetfileHeader,fileHeader,retOpstr,nRegReservedIndex, nLastOffset


# def output_ralf_moduleFile(module_inst, modName):
#     out_ralf_file_Name = './' + modName+'.ralf'
#     with open(out_ralf_file_Name, 'w+') as out_file:
#         fileHeader = """# Autor: Auto generate by python From module excel\n
# # Version: 0.0.2 X
# # Description : struct define for module \n
# # Waring: Do NOT Modify it !

# """

#         nRegData_size = int(module_inst.data_width/8)
#         file_body_str = f'block {modName} ' + \
#             ' {\n\tbytes '+f'{nRegData_size};\n'

#         # 定义module的结构体
#         last_offset = 0
#         nRegReservedIndex = 0

#         bRegGroup = False
#         group_dim = 0
#         group_size = 0
#         group_startPos = 0
#         group_name = ''

#         # print('module data_width: {0}'.format(module_inst.data_width))
#         group_index = -1
#         for reg in module_inst.reg_list:
#             # if reg.bVirtual:
#             #     continue
#             reg_offset = reg.offset
#             # if reg_offset != last_offset:
#             #     # 增加占位
#             #     nRerived = (reg_offset-last_offset) / nRegData_size
#             #     n = 0
#             #     while n < nRerived:
#             #         if bRegGroup:
#             #             file_body_str += f'\tregister reg_reserved{nRegReservedIndex}' + \
#             #                 ' {\n\t\tbytes '+f'{nRegData_size};\n'+'\t}\n'
#             #         nRegReservedIndex += 1
#             #         n += 1
#             if reg.bGroup_start and reg.group_size and reg.group_dim:
#                 bRegGroup = True
#                 group_index = 0
#                 group_startPos = reg_offset
#                 group_size = reg.group_size
#                 group_dim = reg.group_dim
#                 # group_name += reg.reg_name
#                 if reg.bGroup_stop:
#                     file_body_str += f'\tregister {reg.reg_name}[{group_dim}] @{reg.offset} + {group_size}'+' {\n'
#                     bRegGroup = False
#                 else:
#                     file_body_str += f'\tregfile {reg.group_name}[{group_dim}] @{reg.offset} + {group_size}'+' {\n'
#                     file_body_str += f'\t\tregister {reg.reg_name} @{reg.offset-group_startPos}'+' {\n' + \
#                         '\t\tbytes '+f'{nRegData_size};\n'
#             str_Tab = ''
#             if bRegGroup:
#                 str_Tab = '\t'
#                 if reg.group_index != 0:
#                     file_body_str += f'\t\tregister {reg.reg_name} @{reg.offset-group_startPos}'+' {\n' + \
#                         ' \t\tbytes '+f'{nRegData_size};\n'
#             else:
#                 if reg.group_index != 0:
#                     if reg.bVirtual:
#                         file_body_str += f'\tvirtual register  {reg.reg_name} '+' {\n'
#                     else:
#                         file_body_str += f'\tregister  {reg.reg_name} @{reg.offset}'+' {\n'
#                     file_body_str += F'\t\tbytes {nRegData_size};\n'

#             # last_offset = reg_offset + nRegData_size

#             # print('last_offset: is {0}'.format(last_offset))
#             field_count = reg.field_count()
#             if field_count:
#                 # if bRegGroup:
#                 #     file_body_str += f'\t\tregister {reg.reg_name} @{reg.offset-group_startPos}'+' {\n' + \
#                 #     ' \t\tbytes '+f'{nRegData_size};\n'
#                 # else:
#                 #     file_body_str += f'\tregister {reg.reg_name} @{reg.offset}'+' {\n' + \
#                 #     ' \t\tbytes '+f'{nRegData_size};\n'
#                 # nFieldReservedIndex = 0

#                 field_index = field_count-1
#                 # field_bitPos = 0
#                 while field_index != -1:
#                     fd = reg.field_list[field_index]
#                     # if fd.start_bit != field_bitPos:
#                     #     # 需要补齐field
#                     #     if bRegGroup:
#                     #         file_body_str += '\t'
#                     #     file_body_str += f'\t\tfield fd_reserved{nFieldReservedIndex} '+' {\n\t\t\t'
#                     #     if bRegGroup:
#                     #         file_body_str += '\t'
#                     #     file_body_str += f'bits: {fd.start_bit-field_bitPos} ;\n'
#                     #     if bRegGroup:
#                     #         file_body_str += '\t'
#                     #     file_body_str += '\t\t}'
#                     #     nFieldReservedIndex += 1
#                     bReserved = False
#                     field_bitPos = fd.end_bit+1
#                     fd.field_comments = fd.field_comments.replace(
#                         '\n', ' ').replace('\r', ' ')
#                     nBitWid = field_bitPos-fd.start_bit
#                     if fd.field_name == 'reserved':
#                         # fd.field_name = f'reserved{nFieldReservedIndex}'
#                         # nFieldReservedIndex += 1
#                         bReserved = True
#                     if not bReserved:
#                         file_body_str += f'{str_Tab}\t\tfield fd_{fd.field_name} @{fd.start_bit}'+' {\n'
#                         file_body_str += f'{str_Tab}\t\t\tbits {nBitWid} ;\n'
#                         file_body_str += f'{str_Tab}\t\t\treset {fd.defaultValue} ;\n'
#                         if fd.attribute:
#                             file_body_str += f'{str_Tab}\t\t\taccess {fd.attribute.lower()} ;\n'
#                         if fd.field_enumstr:
#                             # print(fd.field_enumstr)
#                             # b_fd_enum = True
#                             enum_lst = fd.field_enumstr.splitlines()
#                             field_enum_str = f'{str_Tab}\t\t\tenum '+' {\n'
#                             b_emFirstitem = True
#                             for em in enum_lst:
#                                 # print(em)
#                                 em_val = em.replace(',', '')
#                                 em_val = em_val.strip()
#                                 (em_item_name, str,
#                                  em_item_value) = em_val.partition('=')
#                                 em_item_name = em_item_name.strip()
#                                 em_item_value = em_item_value.strip().upper()
#                                 if not b_emFirstitem:
#                                     field_enum_str += ',\n'
#                                 if em_item_value and em_item_value.startswith('0X'):
#                                     em_item_value_int = int(em_item_value, 16)
#                                     field_enum_str += f'{str_Tab}\t\t\t\t{em_item_name} {str} {em_item_value_int}'
#                                 else:
#                                     field_enum_str += f'{str_Tab}\t\t\t\t{em_item_name} {str} {em_item_value}'
#                                 b_emFirstitem = False
#                                 # file_str
#                             field_enum_str += f'\n{str_Tab}\t\t\t'+'} \n'
#                             file_body_str += field_enum_str
#                         file_body_str += f'{str_Tab}\t\t\tconstraint c_st_{fd.field_name} ' + '{\n\t\t\t'
#                         file_body_str += str_Tab+'}\n'

#                         file_body_str += str_Tab+'\t\t}'
#                         if fd.field_comments:
#                             file_body_str += f'; #{fd.field_comments} \n'
#                         else:
#                             file_body_str += '\n'
#                     field_index -= 1
#                 # if bRegGroup:
#                 #     reg.group_index = group_index
#                 file_body_str += str_Tab+'\t}; ' + f' #{reg.desc} \n'
#             else:
#                 file_body_str += f'{str_Tab}\tregister  {reg.regname} @{reg.offset-group_startPos}'+' {\n'
#                 file_body_str += F'{str_Tab}\t\tbytes {nRegData_size};\n'
#                 file_body_str += str_Tab+'\t\tfield reserved {\n'
#                 file_body_str += str_Tab+'\t\t\tbits 4 ;\n'
#                 file_body_str += str_Tab+'\t\t}\n'
#                 file_body_str += str_Tab+'\t}'+f';  #{reg.desc}\n'

#             if bRegGroup and reg.bGroup_stop:
#                 # if not reg.bGroup_start:
#                 #     group_name += '__'+reg.reg_name
#                 #     # 需要修改该group的其他reg的groupName
#                 # nRerived = (group_size-group_startPos) / nRegData_size
#                 # n = 0
#                 # while n < nRerived:
#                 #     file_body_str += f'\tregister  reg_reserved{nRegReservedIndex} ' + '{\n'
#                 #     file_body_str += F'\t\tbytes {nRegData_size};\n'
#                 #     file_body_str += '\t}'+f';  #{reg.desc}\n'
#                 #     nRegReservedIndex += 1
#                 #     n += 1
#                 file_body_str += "\t} ; " + \
#                     f'#group_{group_name} [{group_dim}]\n'
#                 # last_offset = group_size*group_dim + group_startPos
#                 bRegGroup = False
#             if bRegGroup:
#                 group_index += 1

#         file_body_str += '\n}; # End of block module '+modName+'\n'
#         out_file.write(fileHeader)
#         out_file.write(file_body_str)
#         out_file.close()
#         return out_ralf_file_Name


# def outModuleFieldDefaultValueCheckCSrc(module_inst_list, modName):
#     # print(modName)
#     dirName = './module_check_defaultvalue/'+modName
#     if not os.path.exists(dirName):
#         os.makedirs(dirName)
#     out_C_file_Name = dirName+'/main.c'
#     with open(out_C_file_Name, 'w+') as out_file:
#         fileHeader = """// Autor: Auto generate by python From module excel\n
# // Version: 0.0.2 X
# // Description : field default value check for module instance \n
# // Waring: Do NOT Modify it !
# // Copyright (C) 2020-2021 CIP United Co. Ltd.  All Rights Reserved.

# #define DEBUG
# //#define INFO
# #define WARNING
# #define NOTICE
# #define ERROR
# #define PASS
# #define FAIL

# #define CHECK_MODULE_FIELD_DEFAULT_VALUE

# #define CHECK_MOUDLE_FIELD_WRITE_VALUE

# #include "log.h"
# #include "pll.h"

# """
#         filebodystr = f'#include "{modName}_reg.h"\n'
#         filebodystr += """
# int main()
# {
#     //printf("enter main.\\n");
#     uAptiv_clk_init();
# """
#         mod_inst_name = modName.upper()
#         mod_inst = module_inst_list[0]
#         filebodystr += f'\tprintf("After clock switch, Now Check Module: {modName}.\\n");\n'
#         if mod_inst.data_width > 32:
#             filebodystr += f'\tuint64_t nRegFdVal = 0;\n'
#         else:
#             filebodystr += f'\tunsigned int nRegFdVal = 0;\n'
#         mod_count = len(module_inst_list)

#         filebodystr += f'\n\tunsigned int nTotalErr = 0;\n'
#         if mod_count > 1:
#             filebodystr += f'\tst_module_info_{modName} * module_inst[{mod_count}] = ' + \
#                 '{'+f'{mod_inst_name}_0'
#             for i in range(1, mod_count):
#                 filebodystr += f'\n\t\t,{mod_inst_name}_{i}'
#             filebodystr += '};\n\n'
#             filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
#             filebodystr += f'\tunsigned int nErrCount_default[{mod_count}] = '+'{0};\n'
#             filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

#             filebodystr += '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
#             filebodystr += f'\tunsigned int nErrCount_wirte[{mod_count}] = '+'{0};\n'
#             filebodystr += '#endif // CHECK_MOUDLE_FIELD_WRITE_VALUE\n\n'

#             filebodystr += f'\tfor(int i = 0; i < {mod_count}; ++i)\n'
#             filebodystr += '\t{\n'
#             filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
#             modinst_var = 'module_inst[i]'
#             errCount_var = 'nErrCount_default[i]'
#             errCount_write_var = 'nErrCount_wirte[i]'
#             str1, str2 = getModuleFdStr(
#                 mod_inst, errCount_var, errCount_write_var, modinst_var)
#             filebodystr += str1
#             filebodystr += '\t\tif(nErrCount_default[i])\n'
#             filebodystr += '\t\t{\n'
#             filebodystr += f'\t\t\tError("Inst_%u def-Vals have [%u] fds NOT Same!\\n", i, nErrCount_default[i]);\n'
#             filebodystr += '\t\t\tnTotalErr += nErrCount_default[i];\n'
#             filebodystr += '\t\t}\n'
#             filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u def-Vals are OK!\\n", i);\n'
#             filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

#             filebodystr += str2
#             filebodystr += '\t\tif(nErrCount_wirte[i])\n'
#             filebodystr += '\t\t{\n'
#             filebodystr += f'\t\t\tError("Inst_%u write-Vals have [%u] fds NOT Same!\\n", i, nErrCount_wirte[i]);\n'
#             filebodystr += '\t\t\tnTotalErr += nErrCount_wirte[i];\n'
#             filebodystr += '\t\t}\n'
#             filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u write-Vals are OK!\\n", i);\n'
#             filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'
#             filebodystr += '\t}\n'
#         elif mod_count == 1:
#             filebodystr += f'\tst_module_info_{modName} * module_inst = {mod_inst_name}_0 ;\n'
#             filebodystr += '#ifdef CHECK_MODULE_FIELD_DEFAULT_VALUE\n'
#             filebodystr += f'\tunsigned int nErrCount_default = 0;\n'
#             filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

#             filebodystr += '#ifdef CHECK_MOUDLE_FIELD_WRITE_VALUE\n'
#             filebodystr += f'\tunsigned int nErrCount_wirte = 0;\n'
#             filebodystr += '#endif // CHECK_MOUDLE_FIELD_WRITE_VALUE\n\n'

#             modinst_var = 'module_inst'
#             errCount_var = 'nErrCount_default'
#             errCount_write_var = 'nErrCount_wirte'
#             # filebodystr += getModuleFdStr(mod_inst,
#             #                               errCount_var, modinst_var, False)
#             str1, str2 = getModuleFdStr(
#                 mod_inst, errCount_var, errCount_write_var, modinst_var, False)

#             filebodystr += str1

#             filebodystr += '\t\tif(nErrCount_default)\n'
#             filebodystr += '\t\t{\n'
#             filebodystr += f'\t\t\tError("Inst_%u def-Vals have [%u] fds NOT Same!\\n", i, nErrCount_default);\n'
#             filebodystr += '\t\t\tnTotalErr += nErrCount_default;\n'
#             filebodystr += '\t\t}\n'
#             filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u def-Vals are OK!\\n", i);\n'
#             filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

#             filebodystr += str2
#             filebodystr += '\t\tif(nErrCount_wirte)\n'
#             filebodystr += '\t\t{\n'
#             filebodystr += f'\t\t\tError("Inst_%u write-Vals have [%u] fds NOT Same!\\n", i, nErrCount_wirte);\n'
#             filebodystr += '\t\t\tnTotalErr += nErrCount_wirte;\n'
#             filebodystr += '\t\t}\n'
#             filebodystr += f'\t\telse\n\t\t\tNotice("Inst_%u write-Vals are OK!\\n", i);\n'
#             filebodystr += '#endif // CHECK_MODULE_FIELD_DEFAULT_VALUE\n\n'

#         # i = 0
#         # for module_inst in module_inst_list:
#         #     # print(module_inst.module_info_str())
#         #     modinst_var = f'{module_inst.module_name.upper()}_{i}'
#         #     group_dim = 0
#         #     for reg in module_inst.reg_list:
#         #         if reg.bVirtual:
#         #             continue
#         #         if reg.bGroup_start and reg.group_dim:
#         #             group_dim = reg.group_dim
#         #         if reg.group_index >= 0 and reg.group_name:
#         #             for g_i in range(0, group_dim):
#         #                 for fd in reg.field_list:
#         #                     if fd.field_name.startswith('reserved'):
#         #                         continue
#         #                     reg_fd_var = f'{reg.reg_name}.fd_{fd.field_name}'
#         #                     fd_var = f'{reg.group_name}[{g_i}].{reg_fd_var}'
#         #                     module_fd_var = f'{modinst_var}->{reg.group_name}[{g_i}].st_reg_{reg_fd_var}'
#         #                     filebodystr += f'\tnRegFdVal = {module_fd_var};\n'
#         #                     filebodystr += f'\tif(nRegFdVal != {fd.defaultValue})\n'
#         #                     filebodystr += '\t{\n'
#         #                     filebodystr += f'\t\tError("Module # FD: {modinst_var} # {fd_var} Value [0x%X] is not same! \\n",nRegFdVal);\n'
#         #                     filebodystr += f'\t\t++nErrCount[{i}];\n'
#         #                     filebodystr += '\t}\n'
#         #                     filebodystr += f'\telse\n\t\tInfo("Module # FD: {modinst_var} # {fd_var} Value is same. \\n");\n'
#         #         else:
#         #             for fd in reg.field_list:
#         #                 if fd.field_name.startswith('reserved'):
#         #                     continue
#         #                 fd_var = f'{reg.reg_name}.fd_{fd.field_name}'
#         #                 module_fd_var = f'{modinst_var}->st_reg_{fd_var}'
#         #                 filebodystr += f'\tnRegFdVal = {module_fd_var};\n'
#         #                 filebodystr += f'\tif(nRegFdVal != {fd.defaultValue})\n'
#         #                 filebodystr += '\t{\n'
#         #                 filebodystr += f'\t\tError("Module # FD: {modinst_var} # {fd_var} Value [0x%X] is not same! \\n",nRegFdVal);\n'
#         #                 filebodystr += f'\t\t++nErrCount[{i}];\n'
#         #                 filebodystr += '\t}\n'
#         #                 filebodystr += f'\telse\n\t\tInfo("Module # FD:: {modinst_var} # {fd_var} Value is same. \\n");\n'
#         #         if reg.bGroup_stop:
#         #             group_dim = 0
#         #     i += 1

#         # filebodystr += f'\n\tunsigned int nTotalErr = 0;\n'
#         # filebodystr +=f'\tfor(int i = 0; i < {mod_count}; ++i)\n'
#         # filebodystr +='\t{\n'
#         # filebodystr += '\t\tif(nErrCount[i])\n'
#         # filebodystr +='\t\t{\n'
#         # filebodystr += f'\t\t\tError("{mod_inst_name}_%u default values have [%u] fields NOT Same!\\n",i,nErrCount[i]);\n'
#         # filebodystr += '\t\t\tnTotalErr += nErrCount[i];\n'
#         # filebodystr +='\t\t}\n'
#         # filebodystr += f'\t\telse\n\t\t\tInfo("{mod_inst_name}_%u default values are All Same!\\n",i);\n'
#         # filebodystr +='\t}\n'
#         filebodystr += f'\tif(nTotalErr == 0)\n'
#         filebodystr += f'\t\tPass("{modName} Vals OK!\\n");\n'
#         filebodystr += f'\telse\n\t\tFail("{modName} Vals Not OK!\\n");\n'
#         filebodystr += '\n\treturn 0;\n}\n'
#         out_file.write(fileHeader)
#         out_file.write(filebodystr)
#         out_file.close()

#         return out_C_file_Name


# def fieldWriteChk_func(errCount_Write_var, str_Tab, fd_var, module_fd_var, strfdMask):
    # fdWriteCheckstr=''
    # fdWriteCheckstr += f'{str_Tab}\t{module_fd_var} = {strfdMask};\n'
    # fdWriteCheckstr += f'{str_Tab}\tnRegFdVal = {module_fd_var};\n'
    # fdWriteCheckstr += f'{str_Tab}\tif({module_fd_var} != {strfdMask})\n{str_Tab}'
    # fdWriteCheckstr += '\t{\n'
    # fdWriteCheckstr += f'{str_Tab}\t\tError("Inst_%u # {fd_var}  [0x%X] NOt same as Write [{strfdMask}]! \\n", i, nRegFdVal);\n'
    # fdWriteCheckstr += f'{str_Tab}\t\t++{errCount_Write_var};\n{str_Tab}'
    # fdWriteCheckstr += '\t}\n'
    # return fdWriteCheckstr


def checkDeviceSheet(ws:worksheet):
    name = ws['B3'].value
    st_dev = St_Device(name)
    vendor = ws['B2'].value
    if vendor:
        st_dev.vendor = vendor 
    version = ws['B4'].value
    if version:
        st_dev.version = version 
    series = ws['B5'].value 
    if series:
        st_dev.series = series
    width = ws['B6'].value
    if width:
        st_dev.width = width
    desc = ws['B7'].value
    if desc: 
        st_dev.description = desc

    row = 9
    mem_row_pos = cpu_row_pos = 0
    nRows = ws.max_row
    maxCols = ws.max_column
    while row <= nRows:
        val = ws.cell(row, 1).value
        if val == 'cpus:':
            cpu_row_pos = row
        elif val == 'memories:':
            mem_row_pos = row
        row += 1

    if cpu_row_pos > 0:
        end_row_pos = nRows
        if mem_row_pos > cpu_row_pos+2:
            end_row_pos = mem_row_pos-1
        if end_row_pos > cpu_row_pos+2:
            #先读出CPU的各列对应的字段
            row_cpu_header =cpu_row_pos+1
            cpu_col_dict = {}
            for col_i in range(1,maxCols):
                val = ws.cell(row_cpu_header,col_i).value
                if val:
                    cpu_col_dict[val] = col_i
            # print(cpu_col_dict)
            for row_i in range(cpu_row_pos+2, end_row_pos):
                cpu_name = ws.cell(row_i, cpu_col_dict['name']).value
                revision = ws.cell(row_i, cpu_col_dict['revision']).value
                endian = ws.cell(row_i, cpu_col_dict['endian']).value
                srs = ws.cell(row_i, cpu_col_dict['srsPresent']).value
                mpu = ws.cell(row_i, cpu_col_dict['mpuPresent']).value
                fpu = ws.cell(row_i, cpu_col_dict['fpuPresent']).value
                dsp = ws.cell(row_i, cpu_col_dict['dspPresent']).value
                icache = ws.cell(row_i,cpu_col_dict['icachePresent']).value
                dcache = ws.cell(row_i,cpu_col_dict['dcachePresent']).value
                mmu = ws.cell(row_i,cpu_col_dict['mmuPresent']).value
                itcm = ws.cell(row_i,cpu_col_dict['itcmPresent']).value
                ditcm = ws.cell(row_i,cpu_col_dict['dtcmPresent']).value
                l2cache = ws.cell(row_i,cpu_col_dict['l2cachePresent']).value
                if cpu_name and revision and endian:
                    st_cpu = St_CPU(cpu_name)
                    derivedFrom = ws.cell(row_i, cpu_col_dict['derivedFrom']).value
                    if derivedFrom: 
                        st_cpu.derivedFrom = derivedFrom
                    st_cpu.revision = revision
                    st_cpu.endian = endian
                    if isinstance(srs,str) :
                        if srs == '1' or srs.upper() == 'TRUE':
                            st_cpu.srsPresent = True
                    elif isinstance(srs,bool):
                        st_cpu.srsPresent = srs
                    if isinstance(mpu,str) :
                        if mpu == '1' or mpu.upper() == 'TRUE':
                            st_cpu.mpuPresent = True
                    elif isinstance(mpu,bool):
                        st_cpu.mpuPresent = mpu
                    if isinstance(fpu,str) :
                        if fpu == '1' or fpu.upper() == 'TRUE':
                            st_cpu.fpuPresent = True
                    elif isinstance(fpu,bool):
                        st_cpu.fpuPresent = fpu
                    if isinstance(dsp,str) :
                        if dsp == '1' or dsp.upper() == 'TRUE':
                            st_cpu.dspPresent = True
                    elif isinstance(dsp,bool):
                        st_cpu.dspPresent = dsp
                    
                    if isinstance(icache,str) :
                        if icache == '1' or icache.upper() == 'TRUE':
                            st_cpu.icachePresent = True
                    elif isinstance(icache,bool):
                        st_cpu.icachePresent = icache

                    if isinstance(dcache,str) :
                        if dcache == '1' or dcache.upper() == 'TRUE':
                            st_cpu.dcache = True
                    elif isinstance(dcache,bool):
                        st_cpu.dcache = dcache

                    if isinstance(mmu,str) :
                        if mmu == '1' or mmu.upper() == 'TRUE':
                            st_cpu.mmu = True
                    elif isinstance(mmu,bool):
                        st_cpu.mmu = mmu

                    if isinstance(itcm,str) :
                        if itcm == '1' or itcm.upper() == 'TRUE':
                            st_cpu.itcmPresent = True
                    elif isinstance(itcm,bool):
                        st_cpu.itcmPresent = itcm

                    if isinstance(ditcm,str) :
                        if ditcm == '1' or ditcm.upper() == 'TRUE':
                            st_cpu.ditcm = True
                    elif isinstance(ditcm,bool):
                        st_cpu.ditcm = ditcm

                    if isinstance(l2cache,str) :
                        if l2cache == '1' or l2cache.upper() == 'TRUE':
                            st_cpu.l2cachePresent = True
                    elif isinstance(l2cache,bool):
                        st_cpu.l2cachePresent = l2cache

                    st_dev.cpus.append(st_cpu)

    if mem_row_pos > 0:
        end_row_pos = nRows
        row_mem_header = mem_row_pos + 1
        mem_col_dict = {}
        for col_i in range(1,maxCols):
            val = ws.cell(row_mem_header,col_i).value
            if val:
                mem_col_dict[val] = col_i
        # print(mem_col_dict)
        if end_row_pos > mem_row_pos+2:
            for row_i in range(mem_row_pos+2, end_row_pos):
                mem_name = ws.cell(row_i, mem_col_dict['name']).value
                addrbase = ws.cell(row_i, mem_col_dict['addressBase']).value
                addrOffset = ws.cell(row_i, mem_col_dict['addressOffset']).value
                size = ws.cell(row_i, mem_col_dict['size']).value
                access = ws.cell(row_i, mem_col_dict['access']).value
                usage = ws.cell(row_i, mem_col_dict['usage']).value
                if mem_name and addrbase and addrOffset and addrOffset and size and access and usage:
                    st_mem = St_Memory(mem_name)
                    derivedFrom = ws.cell(row_i, mem_col_dict['derivedFrom']).value
                    if derivedFrom:
                        st_mem.derivedFrom = derivedFrom
                    st_mem.baseAddress = addrbase
                    st_mem.addressOffset = addrOffset
                    st_mem.size = size
                    st_mem.access = access
                    st_mem.usage = usage
                    processor = ws.cell(row_i, mem_col_dict['processor']).value
                    if processor:
                        st_mem.processor = processor
                    desc = ws.cell(row_i, mem_col_dict['description']).value
                    if desc:
                        st_mem.description = desc
                    st_dev.memories.append(st_mem)

    return st_dev


def dealwith_excel(xls_file:str, outFlag:int = 1):
    # "UART_final_202301010.xls"
    wb = load_workbook(xls_file, data_only=True)
    sheetNames = wb.sheetnames
    st_dev = None
    checkErr = False
    module_file_lst=[]
    for sh_name in sheetNames:
        # 遍历 sheet, preface sheet不处理, module_tpl sheet不处理
        if sh_name in ('preface', 'module_tpl'):
            continue
        elif sh_name == 'device':
            # 先读取device sheet
            ws = wb[sh_name]
            st_dev = checkDeviceSheet(ws)
            # print(st_dev.get_inst_str())
        else:
            # 再逐一读取peripheral sheet
            ws = wb[sh_name]
            err, perip_lst,irq_set = checkModuleSheetValue(ws, sh_name)
            if not err:
                checkErr = True
            else:
                if outFlag == 1:
                    mod_file =output_C_moduleFile(perip_lst, sh_name, st_dev.version)
                    module_file_lst.append(mod_file)
                    output_ralf_moduleFile(perip_lst,sh_name,st_dev.version)
                    output_uvm_sv_moduleFile(perip_lst,sh_name,st_dev.version)
                    output_SequenceSv_moduleFile(perip_lst[0:3],sh_name,st_dev.version)
                    output_C_FdValChk_moduleFile(perip_lst[0:3],sh_name)
                st_dev.peripherals[sh_name] = perip_lst
                if irq_set and isinstance(irq_set,set):
                    st_dev.interrupts.update(irq_set)
           

    if checkErr:
        filename = os.path.basename(xls_file)
        # out_mark_xlsx_file = filename.replace('.xlsx', '_errMk.xlsx')
        print("Check Failed. Please review "+ filename + " and fix it.")
        # wb.save(out_mark_xlsx_file)
    else:
        with open('./dev.json', 'w+') as f: 
            f.write(st_dev.toJson())
        out_sys_uvm_sv(st_dev)
        if outFlag == 2:
            #c代码全部输出到一个文件
            output_C_dev_InOneFile(st_dev)
            pass
        elif outFlag == 1:
            # 输出dev文件
            output_C_devFile(st_dev,module_file_lst)
            pass


if __name__ == '__main__':
    # 全路径是为方便在vscode中进行调试
    # file_name = 'D:/workspace/demopy/excel_flow/excel/ahb_cfg_20230925.xlsx'
    file_name = './xy2_mp32daptyxx_DDF 231208.xlsx'
    dealwith_excel(file_name)
